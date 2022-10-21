from datetime import datetime
start_time = datetime.now()

def octant_range_names(mod=4000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 from operator import itemgetter

 os.system('cls')

 time1 =[]  # make a new list
 v1 = []    # make a new list
 u1 = []    # make a new list
 w1 = []    # make a new list
 v2 = []    # make a list for difference 
 u2 = []    # make a list for difference
 w2 = []    # make a list for difference
 oct =[]    # make a list for octant storing

 wb=load_workbook("octant_input.xlsx")
 sheet=wb["Sheet1"] 
 n=sheet.max_row #n is the number of rows of data
 for r in range(2,n+1): 
  time1.append(sheet.cell(row=r,column=1).value) # making of time1 list from given data
  u1.append(sheet.cell(row=r,column=2).value)  # making of u1 list from given data
  v1.append(sheet.cell(row=r,column=3).value)  # making of v1 list from given data
  w1.append(sheet.cell(row=r,column=4).value)  # making of w1 list from given data

 v_mean=np.mean(v1, dtype=np.float64) # calculate the mean of v1
 u_mean=np.mean(u1, dtype=np.float64) # mean of u1
 w_mean=np.mean(w1, dtype=np.float64) # mean of w1
 
 for r in range(2,n+1):
  u2.append(float(sheet.cell(row=r,column=2).value)-u_mean) # pushing the differences 
  v2.append(float(sheet.cell(row=r,column=3).value)-v_mean) # pushing the differences
  w2.append(float(sheet.cell(row=r,column=4).value)-w_mean)	# pushing the differences  

 for i in range(0,n-1):# using this for loop we will get octant
  if ((u2[i]>=0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(1)
  elif((u2[i]>=0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-1)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(2)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-2)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(4)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]<0 )):
   oct.append(-4)    
  elif((u2[i]<0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(3)
  else:
   oct.append(-3) 
     
 c1=0 # overall count of octant 1
 c2=0 # overall count of octant -1
 c3=0 # overall count of octant 2
 c4=0 # overall count of octant -2
 c5=0 # overall count of octant 3
 c6=0 # overall count of octant -3
 c7=0 # overall count of octant 4
 c8=0 # overall count of octant -4	

 for i in range(0,n-1):# using this for loop we will get octant
  if (oct[i]==1):
   c1=c1+1
  elif(oct[i]==-1):
   c2=c2+1
  elif(oct[i]==2):
   c3=c3+1
  elif(oct[i]==-2):
   c4=c4+1
  elif(oct[i]==3):
   c5=c5+1
  elif(oct[i]==-3):
   c6=c6+1    
  elif(oct[i]==4):
   c7=c7+1
  else:
   c8=c8+1
 rank=[] # this list will give the ranks of all octants in differnt intervals
 rank1_id_name=[] # This list will help rank list in storing the data 
 req=[]  # This list will help rank list in storing the data 
 req.append([c1,1])
 req.append([c2,2])
 req.append([c3,3])
 req.append([c4,4])
 req.append([c5,5])
 req.append([c6,6])
 req.append([c7,7])
 req.append([c8,8]) 
 req.sort()
 req_ranks=[0,0,0,0,0,0,0,0] # this will give ranks of different octant in a particular interval
 req_ranks[req[0][1]-1]=8 # Giving rank 8 in sorted list
 req_ranks[req[1][1]-1]=7 # Giving rank 7 in sorted list
 req_ranks[req[2][1]-1]=6 # Giving rank 6 in sorted list
 req_ranks[req[3][1]-1]=5 # Giving rank 5 in sorted list
 req_ranks[req[4][1]-1]=4 # Giving rank 4 in sorted list
 req_ranks[req[5][1]-1]=3 # Giving rank 3 in sorted list
 req_ranks[req[6][1]-1]=2 # Giving rank 2 in sorted list
 req_ranks[req[7][1]-1]=1 # Giving rank 1 in sorted list
 rank.append(req_ranks) # Appending this list in rank list
 rank_id_name=[]
 if req_ranks[0]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[1,"Internal outward Interaction"]
  rank1_id_name.append(rank_id_name)
 elif req_ranks[1]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[-1,"External outward Interaction"]
  rank1_id_name.append(rank_id_name)
 elif req_ranks[2]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[2,"External Ejection"]
  rank1_id_name.append(rank_id_name) 
 elif req_ranks[3]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[-2,"Internal Ejection"]
  rank1_id_name.append(rank_id_name) 
 elif req_ranks[4]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[3,"External inward Interaction"]
  rank1_id_name.append(rank_id_name) 
 elif req_ranks[5]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[-3,"Internal inward Interaction"]
  rank1_id_name.append(rank_id_name) 
 elif req_ranks[6]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[4,"Internal Sweep"]
  rank1_id_name.append(rank_id_name) 
 elif req_ranks[7]==1: # Givng name and id to Rank 1 Octant
  rank_id_name=[-4,"External Sweep"]
  rank1_id_name.append(rank_id_name) 
 
 o1=[] #make a list of octant count in mod range 
 o2=[] #make a list of octant count in mod range
 o3=[] #make a list of octant count in mod range
 o4=[] #make a list of octant count in mod range
 o5=[] #make a list of octant count in mod range
 o6=[] #make a list of octant count in mod range
 o7=[] #make a list of octant count in mod range
 o8=[] #make a list of octant count in mod range

 oct1=0 #making variables for using in function 
 oct2=0 #making variables for using in function
 oct3=0 #making variables for using in function
 oct4=0 #making variables for using in function
 oct5=0 #making variables for using in function
 oct6=0 #making variables for using in function
 oct7=0 #making variables for using in function
 oct8=0 #making variables for using in function

 m= int((n-2)/mod) +1 # m will give no.of interval
 sn1=0 # This will give Count of rank 1 if Octant 1 in different mod intervals
 sn2=0 # This will give Count of rank -1 if Octant 1 in different mod intervals
 sn3=0 # This will give Count of rank 2if Octant 1 in different mod intervals
 sn4=0 # This will give Count of rank -2 if Octant 1 in different mod intervals
 sn5=0 # This will give Count of rank 3 if Octant 1 in different mod intervals
 sn6=0 # This will give Count of rank -3 if Octant 1 in different mod intervals
 sn7=0 # This will give Count of rank 4 if Octant 1 in different mod intervals
 sn8=0 # This will give Count of rank -4 if Octant 1 in different mod intervals
  
 for i in range(m):
  s=mod*i 
  for j in range(mod):	
   if(j+s<n-1):	                                    
    if (oct[j+s]==1):
     oct1=oct1+1 # calculating no. of octant 1 in particular a interval 
    elif(oct[j+s]==-1):
     oct2=oct2+1 # calculating no. of octant -1 in particular a interval 
    elif(oct[j+s]==2):
     oct3=oct3+1 # calculating no. of octant 2 in particular a interval 
    elif(oct[j+s]==-2):
     oct4=oct4+1 # calculating no. of octant -2 in particular a interval 
    elif(oct[j+s]==3):
     oct5=oct5+1 # calculating no. of octant 3 in particular a interval 
    elif(oct[j+s]==-3):
     oct6=oct6+1   # calculating no. of octant -3 in particular a interval  
    elif(oct[j+s]==4):
     oct7=oct7+1 # calculating no. of octant 4 in particular a interval 
    elif(oct[j+s]==-4):
     oct8=oct8+1 # calculating no. of octant -4 in particular a interval 
  o1.append(oct1)
  o2.append(oct2) 
  o3.append(oct3)
  o4.append(oct4)
  o5.append(oct5)
  o6.append(oct6)
  o7.append(oct7)
  o8.append(oct8)
  req=[] # This list is used for storing the rank in a particular interval
  req.append([oct1,1])
  req.append([oct2,2])
  req.append([oct3,3])
  req.append([oct4,4])
  req.append([oct5,5])
  req.append([oct6,6])
  req.append([oct7,7])
  req.append([oct8,8])
  req.sort() # Sorting req with respect to First Coloumn
  req_ranks=[0,0,0,0,0,0,0,0] # this will give ranks of different octant in a particular interval
  req_ranks[req[0][1]-1]=8 # Giving rank 8 in sorted list
  req_ranks[req[1][1]-1]=7 # Giving rank 7 in sorted list
  req_ranks[req[2][1]-1]=6 # Giving rank 6 in sorted list
  req_ranks[req[3][1]-1]=5 # Giving rank 5 in sorted list
  req_ranks[req[4][1]-1]=4 # Giving rank 4 in sorted list
  req_ranks[req[5][1]-1]=3 # Giving rank 3 in sorted list
  req_ranks[req[6][1]-1]=2 # Giving rank 2 in sorted list
  req_ranks[req[7][1]-1]=1 # Giving rank 1 in sorted list
  rank.append(req_ranks) 
  rank_id_name=[]
  rank_id_name=[]
  if req_ranks[0]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[1,"Internal outward Interaction"]
   rank1_id_name.append(rank_id_name)
   sn1=sn1+1
  elif req_ranks[1]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[-1,"External outward Interaction"]
   rank1_id_name.append(rank_id_name)
   sn2=sn2+1
  elif req_ranks[2]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[2,"External Ejection"]
   rank1_id_name.append(rank_id_name)
   sn3=sn3+1 
  elif req_ranks[3]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[-2,"Internal Ejection"]
   rank1_id_name.append(rank_id_name)
   sn4=sn4+1 
  elif req_ranks[4]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[3,"External inward Interaction"]
   rank1_id_name.append(rank_id_name)
   sn5=sn5+1 
  elif req_ranks[5]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[-3,"Internal inward Interaction"]
   rank1_id_name.append(rank_id_name)
   sn6=sn6+1 
  elif req_ranks[6]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[4,"Internal Sweep"]
   rank1_id_name.append(rank_id_name)
   sn7=sn7+1 
  elif req_ranks[7]==1: # Givng name and id to Rank 1 Octant
   rank_id_name=[-4,"External Sweep"]
   rank1_id_name.append(rank_id_name)
   sn8=sn8+1 
 
  oct1=0
  oct2=0
  oct3=0
  oct4=0
  oct5=0
  oct6=0
  oct7=0
  oct8=0
      # writing of data in sheet1 
 from openpyxl import Workbook 
 book=Workbook()
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4","Rank of 1","Rank of -1","Rank of 2","Rank of -2","Rank of 3","Rank of -3","Rank of 4","Rank of -4","Rank1 Octant ID","Rank1 Octant Name"]) # Printing of heading row
 j=0
 a=0
 b=0
 for x in range(n-2):
  if(x==0):
   rows.append([time1[x],u1[x],v1[x],w1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],oct[x],"","Overall count",c1,c2,c3,c4,c5,c6,c7,c8,rank[x][0],rank[x][1],rank[x][2],rank[x][3],rank[x][4],rank[x][5],rank[x][6],rank[x][7],rank1_id_name[x][0],rank1_id_name[x][1]])
  elif(x==1):
   s="mod "+str(mod)		
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"User input",s,"","","","","","","",""])
  elif(x>=2 and x<2+m):
   if(x==1+m):# it will work only if x=1+m
    z=j*mod 	
    y=(j+1)*mod
    s=str(z)+"-"+str(n-2)		 
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2],rank[x-1][0],rank[x-1][1],rank[x-1][2],rank[x-1][3],rank[x-1][4],rank[x-1][5],rank[x-1][6],rank[x-1][7],rank1_id_name[x-1][0],rank1_id_name[x-1][1]])	 
    j=j+1
   else:
    z=j*mod	
    y=(j+1)*mod-1
    s=str(z)+"-"+str(y)		 
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2],rank[x-1][0],rank[x-1][1],rank[x-1][2],rank[x-1][3],rank[x-1][4],rank[x-1][5],rank[x-1][6],rank[x-1][7],rank1_id_name[x-1][0],rank1_id_name[x-1][1]])	 
    j=j+1
  elif x==m+5: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","Octant ID","Octant Name","Count of Rank 1 Mod values","","","",""])
  elif x==m+6: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","1","Internal Outward Interaction",sn1,"","","",""])
  elif x==m+7: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","-1","External Outward Interaction",sn2,"","","",""])
  elif x==m+8: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","2","External Ejection",sn3,"","","",""])
  elif x==m+9: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","-2","Internal Ejection",sn4,"","","",""])
  elif x==m+10: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","3","External inward Interaction",sn5,"","","",""])
  elif x==m+11: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","-3","Internal inward Interaction",sn6,"","","",""])
  elif x==m+12: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","4","Internal Sweep",sn7,"","","",""]) 
  elif x==m+13: # Printing the table of count of diffent rank 1 octants
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","-4","External Sweep",sn8,"","","",""])     
  else: # Printing the else data
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","","","","","","",""])

 for i in rows:
  sheet.append(i)
 book.save("octant_output_ranking_excel.xlsx")
      
 octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

mod=5000 
octant_range_names(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
