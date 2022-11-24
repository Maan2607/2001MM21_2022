def scorecard():
	pass
#importing required libraries
import openpyxl  
import pandas as pd
import os

from datetime import datetime
start_time = datetime.now()

#Reading files of innings of Pakistan and India
Indian_inn = open("india_inns2.txt","r+") 
Pakista_inn = open("Pak_inns1.txt","r+") 
teams = open("teams.txt","r+")
Team_squads = teams.readlines()

pak_team = Team_squads[0]
pak_players = pak_team[23:-1:].split(",")

India_team = Team_squads[2]
Ind_players = India_team[20:-1:].split(",")


lst_ind=Indian_inn.readlines() 
for i in lst_ind:
    if i=='\n':
        lst_ind.remove(i)
      

lst_pak=Pakista_inn.readlines() 
for i in lst_pak:
    if i=='\n':
        lst_pak.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active

# batting [runs,ball,4s,6s,sr]
# bowling [over,medan,runs,Wickets, NB, WD, ECO]
ind_fow=0                         #indian fall of wickets
pak_fow=0                         # pakistan fall of wickets
out_pb={}                         # Out batter of pakistan
ind_bowlers={}
ind_batter={}

pak_bats={}
pak_bowlers={}
pak_byes=0
pak_bt=0             #total runs given by pak bowlers
for l in lst_pak:
    x=l.index(".")
    over_pak=l[0:x+2]
    temp=l[x+2::].split(",")
    running_ball=temp[0].split("to") #0 2

    if f"{running_ball[0].strip()}" not in ind_bowlers.keys() :
        ind_bowlers[f"{running_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:                        # run count for byes
        if "FOUR" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "1 run" in temp[2]:
            pak_byes+=1
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "2 runs" in temp[2]:
            pak_byes+=2
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "3 runs" in temp[2]:
            pak_byes+=3
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "4 runs" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "5 runs" in temp[2]:
            pak_byes+=5
            ind_bowlers[f"{running_ball[0].strip()}"][0]+=1

    else:
        ind_bowlers[f"{running_ball[0].strip()}"][0]+=1
    
    if f"{running_ball[1].strip()}" not in pak_bats.keys() and temp[1]!="wide":
        pak_bats[f"{running_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        pak_bats[f"{running_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:                                      #scoring for a out 
        ind_bowlers[f"{running_ball[0].strip()}"][3]+=1
        if "Bowled" in temp[1].split("!!")[0]:                           #if bowled
            out_pb[f"{running_ball[1].strip()}"]=("b" + running_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:                         #if got caught
            w=(temp[1].split("!!")[0]).split("by")
            out_pb[f"{running_ball[1].strip()}"]=("c" + w[1] +" b " + running_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:                             # got out by lbw
            out_pb[f"{running_ball[1].strip()}"]=("lbw  b "+running_ball[0])

    

    if "no run" in temp[1] or "out" in temp[1] :                     # scoring for no run on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=0
        pak_bats[f"{running_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:                                        # scoring for 1 run on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=1
        pak_bats[f"{running_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:                                       # scoring for 2 run on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=2
        pak_bats[f"{running_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:                                       # scoring for 3 run on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=3
        pak_bats[f"{running_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:                                       # scoring for 4 run on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=4
        pak_bats[f"{running_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:                                        # scoring for Four  on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=4
        pak_bats[f"{running_ball[1].strip()}"][0]+=4
        pak_bats[f"{running_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:                                         # scoring for six  on current ball
        ind_bowlers[f"{running_ball[0].strip()}"][2]+=6
        pak_bats[f"{running_ball[1].strip()}"][0]+=6
        pak_bats[f"{running_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:                                         # scoring for wide  on current ball
        if "wides" in temp[1]:
            # print(temp[1][1])
            ind_bowlers[f"{running_ball[0].strip()}"][2]+=int(temp[1][1])
            ind_bowlers[f"{running_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            ind_bowlers[f"{running_ball[0].strip()}"][2]+=1
            ind_bowlers[f"{running_ball[0].strip()}"][5]+=1

for val in pak_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)


ind_bowlers_total=0             # indian batting starts
ind_byes=0

out_ind_bat={}
for l in lst_ind:
    x=l.index(".")
    over_ind=l[0:x+2]

    temp=l[x+2::].split(",")

    running_ball=temp[0].split("to") #0 2
    if f"{running_ball[0].strip()}" not in pak_bowlers.keys() :
        pak_bowlers[f"{running_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:
        if "FOUR" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "1" in temp[2]:
            ind_byes+=1
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "2" in temp[2]:
            ind_byes+=2
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "3" in temp[2]:
            ind_byes+=3
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "4" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
        elif "5" in temp[2]:
            ind_byes+=5
            pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
    else:
        pak_bowlers[f"{running_ball[0].strip()}"][0]+=1
    
    if f"{running_ball[1].strip()}" not in ind_batter.keys() and temp[1]!="wide":
        ind_batter[f"{running_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        ind_batter[f"{running_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][3]+=1
        
        if "Bowled" in temp[1].split("!!")[0]:
            out_ind_bat[f"{running_ball[1].strip()}"]=("b" + running_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            out_ind_bat[f"{running_ball[1].strip()}"]=("c" + w[1] +" b " + running_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_ind_bat[f"{running_ball[1].strip()}"]=("lbw  b "+running_ball[0])

    
    
    if "no run" in temp[1] or "out" in temp[1] :
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=0
        ind_batter[f"{running_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=1
        ind_batter[f"{running_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=2
        ind_batter[f"{running_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=3
        ind_batter[f"{running_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=4
        ind_batter[f"{running_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=4
        ind_batter[f"{running_ball[1].strip()}"][0]+=4
        ind_batter[f"{running_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        pak_bowlers[f"{running_ball[0].strip()}"][2]+=6
        ind_batter[f"{running_ball[1].strip()}"][0]+=6
        ind_batter[f"{running_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pak_bowlers[f"{running_ball[0].strip()}"][2]+=int(temp[1][1])
            pak_bowlers[f"{running_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            pak_bowlers[f"{running_ball[0].strip()}"][2]+=1
            pak_bowlers[f"{running_ball[0].strip()}"][5]+=1


for val in ind_batter.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in pak_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in ind_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

for val in pak_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

for val in ind_bowlers.values(): #economy
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1) 


for val in pak_bowlers.values(): #economy
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1)


# pakistan batting starts
names_pak_batters=[]
for key in pak_bats.keys():
    names_pak_batters.append(key)


for i in range(len(pak_bats)):
    sheet.cell(5+i,1).value = names_pak_batters[i]
    sheet.cell(5+i,5).value = pak_bats[names_pak_batters[i]][0]
    sheet.cell(5+i,6).value = pak_bats[names_pak_batters[i]][1]
    sheet.cell(5+i,7).value = pak_bats[names_pak_batters[i]][2]
    sheet.cell(5+i,8).value = pak_bats[names_pak_batters[i]][3]
    sheet.cell(5+i,9).value = pak_bats[names_pak_batters[i]][4]
    if names_pak_batters[i] not in out_pb:
        sheet.cell(5+i,3).value = "not out"
    else:
        sheet.cell(5+i,3).value=out_pb[names_pak_batters[i]]
 
sheet.cell(3,1).value = "BATTERS"          #printing batting stats
sheet["E3"] = "RUNS"
sheet["F3"] = "BALLS"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "

# india bowling stats

sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER"
sheet["D18"] = "MAIDEN"
sheet["E18"] = "RUNS"
sheet["F18"] = "WICKET"
sheet["G18"] = "NO-BALL"
sheet["H18"] = "WIDE"
sheet["I18"] = "ECONOMY"

Pak_BN=[]
for key in pak_bowlers.keys():
    Pak_BN.append(key)

for i in range(len(pak_bowlers)):                      #printing pakistan bowling stats
    sheet.cell(42+i,1).value = Pak_BN[i]
    sheet.cell(42+i,3).value = pak_bowlers[Pak_BN[i]][0]
    sheet.cell(42+i,4).value = pak_bowlers[Pak_BN[i]][1]
    sheet.cell(42+i,5).value = pak_bowlers[Pak_BN[i]][2]
    sheet.cell(42+i,6).value = pak_bowlers[Pak_BN[i]][3]
    sheet.cell(42+i,7).value = pak_bowlers[Pak_BN[i]][4]
    sheet.cell(42+i,8).value = pak_bowlers[Pak_BN[i]][5]
    sheet.cell(42+i,9).value = pak_bowlers[Pak_BN[i]][6]
    pak_bt+=pak_bowlers[Pak_BN[i]][2]
    ind_fow+=pak_bowlers[Pak_BN[i]][3]

# india batting  stats
sheet.cell(11+len(pak_bats)+len(pak_bowlers),1).value = "# INDIA"
sheet.cell(11+len(pak_bats)+len(pak_bowlers),2).value = " INNINGS"

names_ind_batters=[]
for key in ind_batter.keys():
    names_ind_batters.append(key)


