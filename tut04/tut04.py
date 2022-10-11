#Help https://youtu.be/H37f_x4wAC0
def octant_longest_subsequence_count_with_range():
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time1 = [] # make a new list
 v1 = []    # make a new list
 u1 = []    # make a new list
 w1 = []    # make a new list
 v2 = []    # make a list for difference 
 u2 = []    # make a list for difference
 w2 = []    # make a list for difference
 oct = []   # make a list for octant storing

 wb=load_workbook("input_octant_longest_subsequence_with_range.xlsx")
 sheet=wb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time1.append(sheet.cell(row=r,column=1).value)
  u1.append(sheet.cell(row=r,column=2).value)
  v1.append(sheet.cell(row=r,column=3).value)
  w1.append(sheet.cell(row=r,column=4).value)

 v_mean=np.mean(v1, dtype=np.float64) # calculate the mean of v1
 u_mean=np.mean(u1, dtype=np.float64) # mean of u1
 w_mean=np.mean(w1, dtype=np.float64) # mean of w1
 
 for r in range(2,n+1):
  u2.append(float(sheet.cell(row=r,column=2).value)-u_mean) # pushing the differences 
  v2.append(float(sheet.cell(row=r,column=3).value)-v_mean) # pushing the differences
  w2.append(float(sheet.cell(row=r,column=4).value)-w_mean)	# pushing the differences  

 for i in range(0,n-1):# using this for loop we will get octant
  if ((u2[i]>=0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(1)
  elif((u2[i]>=0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-1)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(2)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-2)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(4)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]<0 )):
   oct.append(-4)    
  elif((u2[i]<0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(3)
  else:
   oct.append(-3)
 s1=0 # it is used to store the longest subsequence value of octant 1
 s2=0 # it is used to store the longest subsequence value of octant -1
 s3=0 # it is used to store the longest subsequence value of octant 2
 s4=0 # it is used to store the longest subsequence value of octant -2
 s5=0 # it is used to store the longest subsequence value of octant 3
 s6=0 # it is used to store the longest subsequence value of octant -3
 s7=0 # it is used to store the longest subsequence value of octant 4
 s8=0 # it is used to store the longest subsequence value of octant -4

 a=0 # It will give the longest subsequence length for octant 1
 b=0 # It will give the longest subsequence length for octant -1
 c=0 # It will give the longest subsequence length for octant 2
 d=0 # It will give the longest subsequence length for octant -2
 e=0 # It will give the longest subsequence length for octant 3
 f=0 # It will give the longest subsequence length for octant -3
 g=0 # It will give the longest subsequence length for octant 4
 h=0 # It will give the longest subsequence length for octant -4

 c1=0 # It will give the count for octant 1
 c2=0 # It will give the count for octant -1
 c3=0 # It will give the count for octant 2
 c4=0 # It will give the count for octant -2
 c5=0 # It will give the count for octant 3
 c6=0 # It will give the count for octant -3
 c7=0 # It will give the count for octant 4
 c8=0 # It will give the count for octant -4
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 1
  if oct[r]==1 :
   s1=s1+1
   if r==n-2:
    if s1>a:
     a=s1
     s1=0
     c1=1
    elif a>s1:
     s1=0
    else:
     s1=0
     c1=c1+1 
  else:
    if s1>a:
     a=s1
     s1=0
     c1=1
    elif a>s1:
     s1=0
    else:
     s1=0
     c1=c1+1    
 cn=0 
 t1=[]
 for r in range(0,n-1):  
  if oct[r]==1:
   if cn==0: 
    ti1=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==a:
   t1.append(ti1); 
   cn=0

 print(t1)  
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -1
  if oct[r]==-1 :
   s2=s2+1
   if r==n-2:
    if s2>b:
     b=s2
     s2=0
     c2=1
    elif b>s2:
     s2=0
    else:
     s2=0
     c2=c2+1 
  else:
    if s2>b:
     b=s2
     s2=0
     c2=1
    elif b>s2:
     s2=0
    else:
     s2=0
     c2=c2+1 

 cn=0 
 t2=[]
 for r in range(0,n-1):  
  if oct[r]==-1:
   if cn==0: 
    ti2=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==b:
   t2.append(ti2); 
   cn=0

 print(t2)  
       

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 2
  if oct[r]==2 :
   s3=s3+1
   if r==n-2:
    if s3>c:
     c=s3
     s3=0
     c3=1
    elif c>s3:
     s3=0
    else:
     s3=0
     c3=c3+1 
  else:
    if s3>c:
     c=s3
     s3=0
     c3=1
    elif c>s3:
     s3=0
    else:
     s3=0
     c3=c3+1    
 
 cn=0 
 t3=[]
 for r in range(0,n-1):  
  if oct[r]==2:
   if cn==0: 
    ti3=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==c:
   t3.append(ti3); 
   cn=0

 print(t3)  
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -2
  if oct[r]==-2 :
   s4=s4+1
   if r==n-2:
    if s4>d:
     d=s4
     s4=0
     c4=1
    elif d>s4:
     s4=0
    else:
     s4=0
     c4=c4+1 
  else:
    if s4>d:
     d=s4
     s4=0
     c4=1 
    elif d>s4:
     s4=0
    else:
     s4=0
     c4=c4+1    

 cn=0 
 t4=[]
 for r in range(0,n-1):  
  if oct[r]==-2:
   if cn==0: 
    ti4=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==d:
   t4.append(ti4); 
   cn=0

 print(t4)  
     
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 3
  if oct[r]==3 :
   s5=s5+1
   if r==n-2:
    if s5>e:
     e=s5
     s5=0
     c5=1
    elif e>s5:
     s5=0
    else:
     s5=0
     c5=c5+1 
  else:
    if s5>e:
     e=s5
     s5=0
     c5=1
    elif e>s5:
     s5=0
    else:
     s5=0
     c5=c5+1    
 
 cn=0 
 t5=[]
 for r in range(0,n-1):  
  if oct[r]==3:
   if cn==0: 
    ti5=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==e:
   t5.append(ti5); 
   cn=0

 print(t5)  
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -3
  if oct[r]==-3 :
   s6=s6+1
   if r==n-2:
    if s6>a:
     f=s6
     s6=0
     c6=1
    elif f>s6:
     s6=0
    else:
     s6=0
     c6=c6+1 
  else:
    if s6>f:
     f=s6
     s6=0
     c6=1
    elif f>s6:
     s6=0
    else:
     s6=0
     c6=c6+1    
 
 cn=0 
 t6=[]
 for r in range(0,n-1):  
  if oct[r]==-3:
   if cn==0: 
    ti6=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==f:
   t6.append(ti6); 
   cn=0

 print(t6)  
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 4
  if oct[r]==4 :
   s7=s7+1
   if r==n-2:
    if s7>g:
     g=s7
     s7=0
     c7=1
    elif g>s7:
     s7=0
    else:
     s7=0
     c7=c7+1 
  else:
    if s7>g:
     g=s7
     s7=0
     c7=1
    elif g>s7:
     s7=0
    else:
     s7=0
     c7=c7+1

 cn=0 
 t7=[]
 for r in range(0,n-1):  
  if oct[r]==4:
   if cn==0: 
    ti7=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==g:
   t7.append(ti7); 
   cn=0

 print(t7)  
 
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -4
  if oct[r]==-4 :
   s8=s8+1
   if r==n-2:
    if s8>a:
     h=s8
     s8=0
     c8=1
    elif h>s8:
     s8=0
    else:
     s8=0
     c8=c8+1 
  else:
    if s8>h:
     h=s8
     s8=0
     c8=1
    elif h>s8:
     s8=0
    else:
     s8=0
     c8=c8+1    
       
 cn=0 
 t8=[]
 for r in range(0,n-1):  
  if oct[r]==-4:
   if cn==0: 
    ti8=time1[r] 
   cn=cn+1
  else:
   cn=0 
  if cn==h:
   t8.append(ti8); 
   cn=0

 print(t8) 
 a1=[]
 a1.append("Octant")
 a1.append("1")
 a1.append("Time")
 for i in range(0,c1):
  a1.append("")
 a1.append("-1")
 a1.append("Time")
 for i in range(0,c2):
  a1.append("")
 a1.append("2")
 a1.append("Time")
 for i in range(0,c3):
  a1.append("")
 a1.append("-2")
 a1.append("Time")
 for i in range(0,c4):
  a1.append("")
 a1.append("3")
 a1.append("Time")
 for i in range(0,c5):
  a1.append("")
 a1.append("-3")
 a1.append("Time")
 for i in range(0,c6):
  a1.append("")
 a1.append("4")
 a1.append("Time")
 for i in range(0,c7):
  a1.append("")
 a1.append("-4")
 a1.append("Time")
 for i in range(0,c8):
  a1.append("")
 v=len(a1)
 b1=[]
 b1.append("Longest Subsequence Length")
 b1.append(a)
 b1.append("From")
 for i in range(0,c1):
  b1.append(t1[i])
 b1.append(b)
 b1.append("From")
 for i in range(0,c2):
  b1.append(t2[i])
 b1.append(c)
 b1.append("From")
 for i in range(0,c3):
  b1.append(t3[i])
 b1.append(d)
 b1.append("From")
 for i in range(0,c4):
  b1.append(t4[i])
 b1.append(e)
 b1.append("From")
 for i in range(0,c5):
  b1.append(t5[i])
 b1.append(f)
 b1.append("From")
 for i in range(0,c6):
  b1.append(t6[i])
 b1.append(g)
 b1.append("From")
 for i in range(0,c7):
  b1.append(t7[i])
 b1.append(h)
 b1.append("From")
 for i in range(0,c8):
  b1.append(t8[i])

 d1=[]
 d1.append("Count")
 d1.append(c1)
 d1.append("To")
 for i in range(0,c1):
  d1.append(t1[i]+ 0.01*(a-1))
 d1.append(c2)
 d1.append("To")
 for i in range(0,c2):
  d1.append(t2[i]+ 0.01*(b-1))
 d1.append(c3)
 d1.append("To")
 for i in range(0,c3):
  d1.append(t3[i]+ 0.01*(c-1))
 d1.append(c4)
 d1.append("To")
 for i in range(0,c4):
  d1.append(t4[i]+ 0.01*(d-1))
 d1.append(c5)
 d1.append("To")
 for i in range(0,c5):
  d1.append(t5[i]+ 0.01*(e-1))
 d1.append(c6)
 d1.append("To")
 for i in range(0,c6):
  d1.append(t6[i]+ 0.01*(f-1))
 d1.append(c7)
 d1.append("To")
 for i in range(0,c7):
  d1.append(t7[i]+ 0.01*(g-1))
 d1.append(c8)
 d1.append("To")
 for i in range(0,c8):
  d1.append(t8[i]+ 0.01*(h-1))
 from openpyxl import Workbook 
 book=Workbook()
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 for x in range(n-2): # appending the data in xlsx file 
  if(x==0):
   rows.append([time1[x],u1[x],v1[x],w1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],oct[x],"","Octant","Longest Susequence Length","Count","","Octant","Longest Susequence Length","Count"])
  elif x==1: # appending the data in xlsx file for octant 1  
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","1",a,c1,"",a1[x],b1[x],d1[x]])
  elif x==2: # appending the data in xlsx file for octant -1 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-1",b,c2,"",a1[x],b1[x],d1[x]])
  elif x==3:  # appending the data in xlsx file for octant 2
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","2",c,c3,"",a1[x],b1[x],d1[x]])
  elif x==4:  # appending the data in xlsx file for octant -2
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-2",d,c4,"",a1[x],b1[x],d1[x]]) 
  elif x==5: # appending the data in xlsx file for octant 3 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","3",e,c5,"",a1[x],b1[x],d1[x]])
  elif x==6: # appending the data in xlsx file for octant -3 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-3",f,c6,"",a1[x],b1[x],d1[x]])
  elif x==7: # appending the data in xlsx file for octant 4 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","4",g,c7,"",a1[x],b1[x],d1[x]])
  elif x==8: # appending the data in xlsx file for octant -4 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-4",h,c8,"",a1[x],b1[x],d1[x]])
  elif x<v:
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","","","",a1[x],b1[x],d1[x]])
  else: # appending the remaining data in xlsx file 
   rows.append([time1[x],u1[x],v1[x], w1[x],"","","",u2[x],v2[x],w2[x],oct[x]])

 for i in rows:
  sheet.append(i)
 book.save("output_octant_longest_subsequence_with_range.xlsx")
 
octant_longest_subsequence_count_with_range()