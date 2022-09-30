from logging import BufferingFormatter


def octact_identification(mod):
    import os
    from openpyxl import load_workbook
    import numpy as np
    os.system('cls')
    time_01 = [] # making a new list
    vn1 = []    # new list
    un1 = []    # new list
    wn1 = []    # new list

    v2 = []    # making a list for difference 
    u2 = []    # list for difference
    w2 = []    # list for difference
    octant = [] # making a list for octant storing
    wb=load_workbook("input_octant_transition_identify.xlsx")
    sheet=wb["sheet1"] 
    n=sheet.max_row
    for r in (2,n+1):
     time_01.append(sheet.cell(row=r,column=1).value)
     un1.append(sheet.cell(row=r,column=2).value)
     vn1.append(sheet.cell(row=r,column=3).value)
     wn1.append(sheet.cell(row=r,column=4).value)

    v_mean=np.mean(vn1, dtype=np.float64) # calculate the mean of vn1
    u_mean=np.mean(un1, dtype=np.float64) # mean of un1
    w_mean=np.mean(wn1, dtype=np.float64) # mean of wn1

    for r in (2,n+1):
     u2.append(float(row[1])-u_mean) # push the differences 
     v2.append(float(row[2])-v_mean) # push the differences
     w2.append(float(row[3])-w_mean)	# push the differences  
            
    for i in range(2,n+1):# we will get octant using this for loop 
     if ((u2[i]>=0) and (v2[i]>=0) and (w2[i]>=0 )):
      octant.append(1)
     elif((u2[i]>=0) and (v2[i]>=0) and (w2[i]<0 )):
      octant.append(-1)
     elif((u2[i]<0) and (v2[i]>=0) and (w2[i]>=0 )):
      octant.append(2)
     elif((u2[i]<0) and (v2[i]>=0) and (w2[i]<0 )):
      octant.append(-2)
     elif((u2[i]>=0) and (v2[i]<0) and (w2[i]>=0 )):
      octant.append(4)
     elif((u2[i]>=0) and (v2[i]<0) and (w2[i]<0 )):
      octant.append(-4)    
     elif((u2[i]<0) and (v2[i]<0) and (w2[i]>=0 )):
      octant.append(3)
     else:
      octant.append(-3) 

    c_oct_1=0 # overall count of octant 1
    c_oct_2=0 # overall count of octant -1
    c_oct_3=0 # overall count of octant 2
    c_oct_4=0 # overall count of octant -2
    c_oct_5=0 # overall count of octant 3
    c_oct_6=0 # overall count of octant -3
    c_oct_7=0 # overall count of octant 4
    c_oct_8=0 # overall count of octant -4	 
    for i in range(2,n+1):# we will get octant using this for loop 
     if (octant[i]==1):
      c_oct_1=c_oct_1+1
     elif(octant[i]==-1):
      c_oct_2=c_oct_2+1
     elif(octant[i]==2):
      c_oct_3=c_oct_3+1
     elif(octant[i]==-2):
      c_oct_4=c_oct_4+1
     elif(octant[i]==3):
      c_oct_5=c_oct_5+1
     elif(octant[i]==-3):
      c_oct_6=c_oct_6+1    
     elif(octant[i]==4):
      c_oct_7=c_oct_7+1
     else:
      c_oct_8=c_oct_8+1

    o1=[] # make a list of octant count in mod range 
    o2=[] # make a list of octant count in mod range
    o3=[] # make a list of octant count in mod range
    o4=[] # make a list of octant count in mod range
    o5=[] # make a list of octant count in mod range
    o6=[] # make a list of octant count in mod range
    o7=[] # make a list of octant count in mod range
    o8=[] # make a list of octant count in mod range
    
        
    oct1=0
    oct2=0
    oct3=0
    oct4=0
    oct5=0
    oct6=0
    oct7=0
    oct8=0
    m= int((n-2)/mod) +1
    for i in range(m):
     s=mod*i
      for k in range(mod):	
       if(k+s<n-1):	                                    
         if (octant[k+s]==1):
                        oct1=oct1+1
                    elif(octant[k+s]==-1):
                        oct2=oct2+1
                    elif(octant[k+s]==2):
                        oct3=oct3+1
                    elif(octant[k+s]==-2):
                        oct4=oct4+1
                    elif(octant[k+s]==3):
                        oct5=oct5+1
                    elif(octant[k+s]==-3):
                        oct6=oct6+1    
                    elif(octant[k+s]==4):
                        oct7=oct7+1
                    elif(octant[k+s]==-4):
                        oct8=oct8+1 
            o1.append(oct1)
            o2.append(oct2)
            o3.append(oct3)
            o4.append(oct4)
            o5.append(oct5)
            o6.append(oct6)
            o7.append(oct7)
            o8.append(oct8)
            oct1=0
            oct2=0
            oct3=0
            oct4=0
            oct5=0
            oct6=0
            oct7=0
            oct8=0



             write the code 

             
        from openpyxl import workbook 
        book=workbook()
        sheet= book.active    
        rows=[] 
        rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4"])
   )   
      for i in range(2,n+1):
      if(i==2)
        rows.append([time_01[x],un1[x],vn1[x],wn1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],octant[x],"","Overall count",c_oct_1,c_oct_2,c_oct_3,c_oct_4,c_oct_5,c_oct_6,c_oct_7,c_oct_8])
)
      // same 



    from i in rows:
     sheet.append(i)
    book.save(output file)


      






    with open('octant_output.csv','w',newline="") as file:
        writer=csv.writer(file)
        writer.writerow(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4"])
        k=0
        for x in range(n-1):
            if(x==0):
                writer.writerow([time_01[x],un1[x],vn1[x],wn1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],octant[x],"","Overall count",c_oct_1,c_oct_2,c_oct_3,c_oct_4,c_oct_5,c_oct_6,c_oct_7,c_oct_8])
            elif(x==1):
                s="mod "+str(mod)		
                writer.writerow([time_01[x],un1[x],vn1[x],wn1[x],"","","",u2[x],v2[x],w2[x],octant[x],"User input",s,"","","","","","","",""])
            elif(x>=2 and x<2+m):
                if(x==1+m):
                    z=k*mod 	
                    y=(k+1)*mod
                    s=str(z)+"-"+str(n)		 
                    writer.writerow([time_01[x],un1[x],vn1[x],wn1[x],"","","",u2[x],v2[x],w2[x],octant[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2]])	 
                    k=k+1
                else:
                    z=k*mod	
                    y=(k+1)*mod-1
                    s=str(z)+"-"+str(y)		 
                    writer.writerow([time_01[x],un1[x],vn1[x],wn1[x],"","","",u2[x],v2[x],w2[x],octant[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2]])	 
                    k=k+1
            else:
                writer.writerow([time_01[x],un1[x],vn1[x],wn1[x],"","","",u2[x],v2[x],w2[x],octant[x],"","","","","","","","","",""])			 

mod = 5000
octact_identification(mod)