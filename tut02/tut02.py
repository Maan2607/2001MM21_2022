def octant_transition_count(mod=5000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time1 = [] # make a new list
 v1 = []    # make a new list
 u1 = []    # make a new list
 w1 = []    # make a new list
 v2 = []    # make a list for difference 
 u2 = []    # make a list for difference
 w2 = []    # make a list for difference
 oct = []   # make a list for octant storing

 wb=load_workbook("input_octant_transition_identify.xlsx")
 sheet=wb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time1.append(sheet.cell(row=r,column=1).value)
  u1.append(sheet.cell(row=r,column=2).value)
  v1.append(sheet.cell(row=r,column=3).value)
  w1.append(sheet.cell(row=r,column=4).value)

 v_mean=np.mean(v1, dtype=np.float64) # calculate the mean of v1
 u_mean=np.mean(u1, dtype=np.float64) # mean of u1
 w_mean=np.mean(w1, dtype=np.float64) # mean of w1
 
 for r in range(2,n+1):
  u2.append(float(sheet.cell(row=r,column=2).value)-u_mean) # pushing the differences 
  v2.append(float(sheet.cell(row=r,column=3).value)-v_mean) # pushing the differences
  w2.append(float(sheet.cell(row=r,column=4).value)-w_mean)	# pushing the differences  

 for i in range(0,n-1):# using this for loop we will get octant
  if ((u2[i]>=0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(1)
  elif((u2[i]>=0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-1)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(2)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-2)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(4)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]<0 )):
   oct.append(-4)    
  elif((u2[i]<0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(3)
  else:
   oct.append(-3) 
     
 c1=0 # overall count of octant 1
 c2=0 # overall count of octant -1
 c3=0 # overall count of octant 2
 c4=0 # overall count of octant -2
 c5=0 # overall count of octant 3
 c6=0 # overall count of octant -3
 c7=0 # overall count of octant 4
 c8=0 # overall count of octant -4	

 for i in range(0,n-1):# using this for loop we will get octant
  if (oct[i]==1):
   c1=c1+1
  elif(oct[i]==-1):
   c2=c2+1
  elif(oct[i]==2):
   c3=c3+1
  elif(oct[i]==-2):
   c4=c4+1
  elif(oct[i]==3):
   c5=c5+1
  elif(oct[i]==-3):
   c6=c6+1    
  elif(oct[i]==4):
   c7=c7+1
  else:
   c8=c8+1

 o1=[] #make a list of octant count in mod range 
 o2=[] #make a list of octant count in mod range
 o3=[] #make a list of octant count in mod range
 o4=[] #make a list of octant count in mod range
 o5=[] #make a list of octant count in mod range
 o6=[] #make a list of octant count in mod range
 o7=[] #make a list of octant count in mod range
 o8=[] #make a list of octant count in mod range

 oct1=0 #making variables for using in function 
 oct2=0 #making variables for using in function
 oct3=0 #making variables for using in function
 oct4=0 #making variables for using in function
 oct5=0 #making variables for using in function
 oct6=0 #making variables for using in function
 oct7=0 #making variables for using in function
 oct8=0 #making variables for using in function

 m= int((n-2)/mod) +1 # m will give no.of interval

 for i in range(m):
  s=mod*i
  for j in range(mod):	
   if(j+s<n-1):	                                    
    if (oct[j+s]==1):
     oct1=oct1+1 
    elif(oct[j+s]==-1):
     oct2=oct2+1
    elif(oct[j+s]==2):
     oct3=oct3+1
    elif(oct[j+s]==-2):
     oct4=oct4+1
    elif(oct[j+s]==3):
     oct5=oct5+1
    elif(oct[j+s]==-3):
     oct6=oct6+1    
    elif(oct[j+s]==4):
     oct7=oct7+1
    elif(oct[j+s]==-4):
     oct8=oct8+1 
  o1.append(oct1)
  o2.append(oct2) 
  o3.append(oct3)
  o4.append(oct4)
  o5.append(oct5)
  o6.append(oct6)
  o7.append(oct7)
  o8.append(oct8)
  oct1=0
  oct2=0
  oct3=0
  oct4=0
  oct5=0
  oct6=0
  oct7=0
  oct8=0

 r=[] # it is giving full data that we need to print
 req=[] # it will give the data 2-d matrix of respective interval
 over=[[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8] # This is matirx of overall transition
 for i in range(0,m) :
  p=i*mod
  a1=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a2=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a3=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a4=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a5=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a6=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a7=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
  a8=[0]*8 # These list will store the value each element in 2-d matrix of different intervals 
  for j in range(0,mod):
   if j+p<=n-3:
    if oct[j+p]==1 and oct[j+p+1]==1 :
      a1[0]=a1[0]+1 # This will keep on increasing the value of each transition in different intervals 
      over[0][0]=over[0][0]+1 # This is for overall transitin matrix
    elif oct[j+p]==1 and oct[j+p+1]==-1:
      a1[1]=a1[1]+1  # This will keep on increasing the value of each transition in different intervals  
      over[0][1]=over[0][1]+1 # This is for overall transitin matrix
    elif oct[j+p]==1 and oct[j+p+1]==2:
      a1[2]=a1[2]+1
      over[0][2]=over[0][2]+1
    elif oct[j+p]==1 and oct[j+p+1]==-2:
      a1[3]=a1[3]+1
      over[0][3]=over[0][3]+1
    elif oct[j+p]==1 and oct[j+1+p]==3:
      a1[4]=a1[4]+1 
      over[0][4]=over[0][4]+1
    elif oct[j+p]==1 and oct[j+p+1]==-3:
      a1[5]=a1[5]+1
      over[0][5]=over[0][5]+1
    elif oct[j+p]==1 and oct[j+1+p]==4:
      a1[6]=a1[6]+1
      over[0][6]=over[0][6]+1
    elif oct[j+p]==1 and oct[j+p+1]==-4:
      a1[7]=a1[7]+1
      over[0][7]=over[0][7]+1
  
    elif oct[j+p]==-1 and oct[j+p+1]==1 :
      a2[0]=a2[0]+1
      over[1][0]=over[1][0]+1
    elif oct[j+p]==-1 and oct[j+p+1]==-1:
      a2[1]=a2[1]+1    
      over[1][1]=over[1][1]+1
    elif oct[j+p]==-1 and oct[j+p+1]==2:
      a2[2]=a2[2]+1
      over[1][2]=over[1][2]+1
    elif oct[j+p]==-1 and oct[j+p+1]==-2:
      a2[3]=a2[3]+1
      over[1][3]=over[1][3]+1
    elif oct[j+p]==-1 and oct[j+p+1]==3:
      a2[4]=a2[4]+1
      over[1][4]=over[1][4]+1 
    elif oct[j+p]==-1 and oct[j+1+p]==-3:
      a2[5]=a2[5]+1
      over[1][5]=over[1][5]+1
    elif oct[j+p]==-1 and oct[j+p+1]==4:
      a2[6]=a2[6]+1
      over[1][6]=over[1][6]+1
    elif oct[j+p]==-1 and oct[j+1+p]==-4:
      a2[7]=a2[7]+1
      over[1][7]=over[1][7]+1
    
    elif oct[j+p]==2 and oct[j+p+1]==1 :
      a3[0]=a3[0]+1
      over[2][0]=over[2][0]+1
    elif oct[j+p]==2 and oct[j+p+1]==-1:
      a3[1]=a3[1]+1
      over[2][1]=over[2][1]+1    
    elif oct[j+p]==2 and oct[j+p+1]==2:
      a3[2]=a3[2]+1
      over[2][2]=over[2][2]+1
    elif oct[j+p]==2 and oct[j+p+1]==-2:
      a3[3]=a3[3]+1
      over[2][3]=over[2][3]+1
    elif oct[j+p]==2 and oct[j+p+1]==3:
      a3[4]=a3[4]+1
      over[2][4]=over[2][4]+1 
    elif oct[j+p]==2 and oct[j+1+p]==-3:
      a3[5]=a3[5]+1
      over[2][5]=over[2][5]+1
    elif oct[j+p]==2 and oct[j+p+1]==4:
      a3[6]=a3[6]+1
      over[2][6]=over[2][6]+1
    elif oct[j+p]==2 and oct[j+1+p]==-4:
      a3[7]=a3[7]+1
      over[2][7]=over[2][7]+1

    elif oct[j+p]==-2 and oct[j+p+1]==1 :
      a4[0]=a4[0]+1
      over[3][0]=over[3][0]+1
    elif oct[j+p]==-2 and oct[j+p+1]==-1:
      a4[1]=a4[1]+1
      over[3][1]=over[3][1]+1    
    elif oct[j+p]==-2 and oct[j+p+1]==2:
      a4[2]=a4[2]+1
      over[3][2]=over[3][2]+1
    elif oct[j+p]==-2 and oct[j+p+1]==-2:
      a4[3]=a4[3]+1
      over[3][3]=over[3][3]+1
    elif oct[j+p]==-2 and oct[j+p+1]==3:
      a4[4]=a4[4]+1 
      over[3][4]=over[3][4]+1
    elif oct[j+p]==-2 and oct[j+1+p]==-3:
      a4[5]=a4[5]+1
      over[3][5]=over[3][5]+1
    elif oct[j+p]==-2 and oct[j+p+1]==4:
      a4[6]=a4[6]+1
      over[3][6]=over[3][6]+1
    elif oct[j+p]==-2 and oct[j+1+p]==-4:
      a4[7]=a4[7]+1
      over[3][7]=over[3][7]+1  
  
    elif oct[j+p]==3 and oct[j+p+1]==1 :
      a5[0]=a5[0]+1
      over[4][0]=over[4][0]+1
    elif oct[j+p]==3 and oct[j+p+1]==-1:
      a5[1]=a5[1]+1
      over[4][1]=over[4][1]+1    
    elif oct[j+p]==3 and oct[j+p+1]==2:
      a5[2]=a5[2]+1
      over[4][2]=over[4][2]+1
    elif oct[j+p]==3 and oct[j+p+1]==-2:
      a5[3]=a5[3]+1
      over[4][3]=over[4][3]+1
    elif oct[j+p]==3 and oct[j+p+1]==3:
      a5[4]=a5[4]+1 
      over[4][4]=over[4][4]+1
    elif oct[j+p]==3 and oct[j+1+p]==-3:
      a5[5]=a5[5]+1
      over[4][5]=over[4][5]+1
    elif oct[j+p]==3 and oct[j+p+1]==4:
      a5[6]=a5[6]+1
      over[4][6]=over[4][6]+1
    elif oct[j+p]==3 and oct[j+1+p]==-4:
      a5[7]=a5[7]+1
      over[4][7]=over[4][7]+1 
    
    elif oct[j+p]==-3 and oct[j+p+1]==1 :
      a6[0]=a6[0]+1
      over[5][0]=over[5][0]+1
    elif oct[j+p]==-3 and oct[j+p+1]==-1:
      a6[1]=a6[1]+1    
      over[5][1]=over[5][1]+1
    elif oct[j+p]==-3 and oct[j+p+1]==2:
      a6[2]=a6[2]+1
      over[5][2]=over[5][2]+1
    elif oct[j+p]==-3 and oct[j+p+1]==-2:
      a6[3]=a6[3]+1
      over[5][3]=over[5][3]+1
    elif oct[j+p]==-3 and oct[j+p+1]==3:
      a6[4]=a6[4]+1
      over[5][4]=over[5][4]+1 
    elif oct[j+p]==-3 and oct[j+1+p]==-3:
      a6[5]=a6[5]+1
      over[5][5]=over[5][5]+1
    elif oct[j+p]==-3 and oct[j+p+1]==4:
      a6[6]=a6[6]+1
      over[5][6]=over[5][6]+1
    elif oct[j+p]==-3 and oct[j+1+p]==-4:
      a6[7]=a6[7]+1
      over[5][7]=over[5][7]+1

    elif oct[j+p]==4 and oct[j+p+1]==1 :
      a7[0]=a7[0]+1
      over[6][0]=over[6][0]+1
    elif oct[j+p]==4 and oct[j+p+1]==-1:
      a7[1]=a7[1]+1    
      over[6][1]=over[6][1]+1
    elif oct[j+p]==4 and oct[j+p+1]==2:
      a7[2]=a7[2]+1
      over[6][2]=over[6][2]+1
    elif oct[j+p]==4 and oct[j+p+1]==-2:
      a7[3]=a7[3]+1
      over[6][3]=over[6][3]+1
    elif oct[j+p]==4 and oct[j+p+1]==3:
      a7[4]=a7[4]+1
      over[6][4]=over[6][4]+1
    elif oct[j+p]==4 and oct[j+1+p]==-3:
      a7[5]=a7[5]+1
      over[6][5]=over[6][5]+1
    elif oct[j+p]==4 and oct[j+p+1]==4:
      a7[6]=a7[6]+1
      over[6][6]=over[6][6]+1
    elif oct[j+p]==4 and oct[j+1+p]==-4:
      a7[7]=a7[7]+1 
      over[6][7]=over[6][7]+1
    elif oct[j+p]==-4 and oct[j+p+1]==1 :
      a8[0]=a8[0]+1
      over[7][0]=over[7][0]+1
    elif oct[j+p]==-4 and oct[j+p+1]==-1:
      a8[1]=a8[1]+1 
      over[7][1]=over[7][1]+1  
    elif oct[j+p]==-4 and oct[j+p+1]==2:
      a8[2]=a8[2]+1
      over[7][2]=over[7][2]+1
    elif oct[j+p]==-4 and oct[j+p+1]==-2:
      a8[3]=a8[3]+1
      over[7][3]=over[7][3]+1
    elif oct[j+p]==-4 and oct[j+p+1]==3:
      a8[4]=a8[4]+1
      over[7][4]=over[7][4]+1 
    elif oct[j+p]==-4 and oct[j+1+p]==-3:
      a8[5]=a8[5]+1
      over[7][5]=over[7][5]+1
    elif oct[j+p]==-4 and oct[j+p+1]==4:
      a8[6]=a8[6]+1
      over[7][6]=over[7][6]+1
    elif oct[j+p]==-4 and oct[j+1+p]==-4:
      a8[7]=a8[7]+1  
      over[7][7]=over[7][7]+1    
  if i==0 :
    r.append(over)
  req.append(a1)  # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a2) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a3) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a4) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a5) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a6) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a7) # Making a 2-d Matrix req by appending eight 1-d matrix
  req.append(a8) # Making a 2-d Matrix req by appending eight 1-d matrix
  c=req.copy() 
  r.append(c) # appending the req matrix in r
  req.clear() # clearing req matrix so that new data may get store into it
 h = ["+1","-1","+2","-2","+3","-3","+4","-4"]
 
 from openpyxl import Workbook 
 book=Workbook()
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4"])
 j=0
 a=0
 b=0
 for x in range(n-2):
  if(x==0):
   rows.append([time1[x],u1[x],v1[x],w1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],oct[x],"","Overall count",c1,c2,c3,c4,c5,c6,c7,c8])
  elif(x==1):
   s="mod "+str(mod)		
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"User input",s,"","","","","","","",""])
  elif(x>=2 and x<2+m):
   if(x==1+m):# it will work ont=ly if x=1+m
    z=j*mod 	
    y=(j+1)*mod
    s=str(z)+"-"+str(n-2)		 
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2]])	 
    rows.append([time1[x],u1[x],v1[x],w1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],oct[x],"","Verified",c1,c2,c3,c4,c5,c6,c7,c8])
    j=j+1
   else:
    z=j*mod	
    y=(j+1)*mod-1
    s=str(z)+"-"+str(y)		 
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",s,o1[x-2],o2[x-2],o3[x-2],o4[x-2],o5[x-2],o6[x-2],o7[x-2],o8[x-2]])	 
    j=j+1
  elif (x-(2+m))%9==0 and x<9*(m+1)+2+m:
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","","","","","","","",""])
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","","","","","","","",""])
    if x ==2+m:
     rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","Overall transition Count","","","","","","","",""])
     rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","To","","","","","","",""])
    else:
     rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","Mod transition Count","","","","","","","",""])
     rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",str(((x-(2+m))//9-1)*mod)+"-"+str(np.minimum(((x-(2+m))//9)*mod,n-2)),"To","","","","","","",""])
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","Count","+1","-1","+2","-2","+3","-3","+4","-4"])
  elif x<9*(m+1)+(2+m):
   if (x-(2+m))%9==1:
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"From",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])	
   else:
    rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])
   a=a+1
   if a==8:
    a=0
    b=b+1
  else:
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","","","","","","","",""])

 for i in rows:
  sheet.append(i)
 book.save("output_octant_transition_identify.xlsx")

mod=2000
octant_transition_count(mod)