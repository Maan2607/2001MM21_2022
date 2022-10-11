def octant_longest_subsequence_count():
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time1 = [] # make a new list
 v1 = []    # make a new list
 u1 = []    # make a new list
 w1 = []    # make a new list
 v2 = []    # make a list for difference 
 u2 = []    # make a list for difference
 w2 = []    # make a list for difference
 oct = []   # make a list for octant storing

 wb=load_workbook("input_octant_longest_subsequence.xlsx")
 sheet=wb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time1.append(sheet.cell(row=r,column=1).value)
  u1.append(sheet.cell(row=r,column=2).value)
  v1.append(sheet.cell(row=r,column=3).value)
  w1.append(sheet.cell(row=r,column=4).value)

 v_mean=np.mean(v1, dtype=np.float64) # calculate the mean of v1
 u_mean=np.mean(u1, dtype=np.float64) # mean of u1
 w_mean=np.mean(w1, dtype=np.float64) # mean of w1
 
 for r in range(2,n+1):
  u2.append(float(sheet.cell(row=r,column=2).value)-u_mean) # pushing the differences 
  v2.append(float(sheet.cell(row=r,column=3).value)-v_mean) # pushing the differences
  w2.append(float(sheet.cell(row=r,column=4).value)-w_mean)	# pushing the differences  

 for i in range(0,n-1):# using this for loop we will get octant
  if ((u2[i]>=0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(1)
  elif((u2[i]>=0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-1)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]>=0 )):
   oct.append(2)
  elif((u2[i]<0) and (v2[i]>=0) and (w2[i]<0 )):
   oct.append(-2)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(4)
  elif((u2[i]>=0) and (v2[i]<0) and (w2[i]<0 )):
   oct.append(-4)    
  elif((u2[i]<0) and (v2[i]<0) and (w2[i]>=0 )):
   oct.append(3)
  else:
   oct.append(-3)
 s1=0 # it is used to store the  longest subsequence value of octant 1
 s2=0 # it is used to store the longest subsequence value of octant -1
 s3=0 # it is used to store the longest subsequence value of octant 2
 s4=0 # it is used to store the longest subsequence value of octant -2
 s5=0 # it is used to store the longest subsequence value of octant 3
 s6=0 # it is used to store the longest subsequence value of octant -3
 s7=0 # it is used to store the longest subsequence value of octant 4
 s8=0 # it is used to store the longest subsequence value of octant -4

 a=0 # It will give the longest subsequence length for octant 1
 b=0 # It will give the longest subsequence length for octant -1
 c=0 # It will give the longest subsequence length for octant 2
 d=0 # It will give the longest subsequence length for octant -2
 e=0 # It will give the longest subsequence length for octant 3
 f=0 # It will give the longest subsequence length for octant -3
 g=0 # It will give the longest subsequence length for octant 4
 h=0 # It will give the longest subsequence length for octant -4

 c1=0 # It will give the count for octant 1
 c2=0 # It will give the count for octant -1
 c3=0 # It will give the count for octant 2
 c4=0 # It will give the count for octant -2
 c5=0 # It will give the count for octant 3
 c6=0 # It will give the count for octant -3
 c7=0 # It will give the count for octant 4
 c8=0 # It will give the count for octant -4

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 1
  if oct[r]==1 :
   s1=s1+1
   if r==n-2:
    if s1>a:
     a=s1
     s1=0
     c1=1
    elif a>s1:
     s1=0
    else:
     s1=0
     c1=c1+1 
  else:
    if s1>a:
     a=s1
     s1=0
     c1=1
    elif a>s1:
     s1=0
    else:
     s1=0
     c1=c1+1    

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -1
  if oct[r]==-1 :
   s2=s2+1
   if r==n-2:
    if s2>b:
     b=s2
     s2=0
     c2=1
    elif b>s2:
     s2=0
    else:
     s2=0
     c2=c2+1 
  else:
    if s2>b:
     b=s2
     s2=0
     c2=1
    elif b>s2:
     s2=0
    else:
     s2=0
     c2=c2+1   

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 2
  if oct[r]==2 :
   s3=s3+1
   if r==n-2:
    if s3>c:
     c=s3
     s3=0
     c3=1
    elif c>s3:
     s3=0
    else:
     s3=0
     c3=c3+1 
  else:
    if s3>c:
     c=s3
     s3=0
     c3=1
    elif c>s3:
     s3=0
    else:
     s3=0
     c3=c3+1    

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -2
  if oct[r]==-2 :
   s4=s4+1
   if r==n-2:
    if s4>d:
     d=s4
     s4=0
     c4=1
    elif d>s4:
     s4=0
    else:
     s4=0
     c4=c4+1 
  else:
    if s4>d:
     d=s4
     s4=0
     c4=1 
    elif d>s4:
     s4=0
    else:
     s4=0
     c4=c4+1    
     
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 3
  if oct[r]==3 :
   s5=s5+1
   if r==n-2:
    if s5>e:
     e=s5
     s5=0
     c5=1
    elif e>s5:
     s5=0
    else:
     s5=0
     c5=c5+1 
  else:
    if s5>e:
     e=s5
     s5=0
     c5=1
    elif e>s5:
     s5=0
    else:
     s5=0
     c5=c5+1    

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -3
  if oct[r]==-3 :
   s6=s6+1
   if r==n-2:
    if s6>a:
     f=s6
     s6=0
     c6=1
    elif f>s6:
     s6=0
    else:
     s6=0
     c6=c6+1 
  else:
    if s6>f:
     f=s6
     s6=0
     c6=1
    elif f>s6:
     s6=0
    else:
     s6=0
     c6=c6+1    

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 4
  if oct[r]==4 :
   s7=s7+1
   if r==n-2:
    if s7>g:
     g=s7
     s7=0
     c7=1
    elif g>s7:
     s7=0
    else:
     s7=0
     c7=c7+1 
  else:
    if s7>g:
     g=s7
     s7=0
     c7=1
    elif g>s7:
     s7=0
    else:
     s7=0
     c7=c7+1

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant -4
  if oct[r]==-4 :
   s8=s8+1
   if r==n-2:
    if s8>a:
     h=s8
     s8=0
     c8=1
    elif h>s8:
     s8=0
    else:
     s8=0
     c8=c8+1 
  else:
    if s8>h:
     h=s8
     s8=0
     c8=1
    elif h>s8:
     s8=0
    else:
     s8=0
     c8=c8+1    
       

 from openpyxl import Workbook 
 book=Workbook()
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 for x in range(n-2): # appending the data in xlsx file 
  if(x==0):
   rows.append([time1[x],u1[x],v1[x],w1[x],u_mean,v_mean,w_mean,u2[x],v2[x],w2[x],oct[x],"","Octant","Longest Susequence Length","Count"])
  elif x==1: # appending the data in xlsx file for octant 1  
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","1",a,c1])
  elif x==2: # appending the data in xlsx file for octant -1 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-1",b,c2])
  elif x==3:  # appending the data in xlsx file for octant 2
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","2",c,c3])
  elif x==4:  # appending the data in xlsx file for octant -2
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-2",d,c4]) 
  elif x==5: # appending the data in xlsx file for octant 3 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","3",e,c5])
  elif x==6: # appending the data in xlsx file for octant -3 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-3",f,c6])
  elif x==7: # appending the data in xlsx file for octant 4 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","4",g,c7])
  elif x==8: # appending the data in xlsx file for octant -4 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x],"","-4",h,c8])
  else: # appending the remaining data in xlsx file 
   rows.append([time1[x],u1[x],v1[x],w1[x],"","","",u2[x],v2[x],w2[x],oct[x]])

 for i in rows:
  sheet.append(i)
 book.save("output_octant_longest_subsequence.xlsx")
   
octant_longest_subsequence_count()