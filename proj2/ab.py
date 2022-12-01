import streamlit as st
from openpyxl.styles.borders import Border, Side
import pandas as pd
import os.path
import pathlib
import shutil
from io import BytesIO


border_them = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
head_comp=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']
def find_oct(a,b,c):                                      
	if(a>0 and b>0 and c>0):    
		return 1
	elif(a>0 and b>0 and c<0):
		return -1
	elif(a<0 and b>0 and c>0):
		return 2
	elif(a<0 and b>0 and c<0):
		return -2
	elif(a<0 and b<0 and c>0):
		return 3
	elif(a<0 and b<0 and c<0):
		return -3
	elif(a>0 and b<0 and c>0):
		return 4
	elif(a>0 and b<0 and c<0):
		return -4

def octant_analysis_single(mod,file):
	from pandas import read_excel
	import openpyxl
	from openpyxl import workbook,load_workbook
	from openpyxl import Workbook
	from openpyxl.styles import PatternFill
	from itertools import repeat

	def workingFile(s,mod):							
		inputFilePath=s									
		df=read_excel(inputFilePath)									
		wb=Workbook()											
		sheet=wb.active													
		
		for i in range(11):										
			sheet.cell(row=2,column=i+1).value=head_comp[i]	
		octant=[]
		# sheet.cell(2,1).value = 'T'
		# sheet.cell(2,2).value = 'U'
		# sheet.cell(2,3).value = 'V'
		# sheet.cell(2,4).value = 'W'
		# sheet.cell(2,5).value = 'U Avg'
		# sheet.cell(2,6).value = 'V Avg'
		# sheet.cell(2,7).value = 'W Avg'
		# sheet.cell(2,8).value = "U'=U-U avg"
		# sheet.cell(2,9).value = "V'=V-V avg"
		# sheet.cell(2,10).value = "W'=W-W avg"
		# sheet.cell(2,11).value = 'Octant'

		u_avg=df['U'].mean()                                                 
		v_avg=df['V'].mean()
		w_avg=df['W'].mean()
		sheet.cell(row=1,column=14).value='Overall Octant Count'
		sheet.cell(row=1,column=45).value='Longest Subsequence Length'
		sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
		sheet['E3'] = u_avg
		sheet['F3'] = v_avg
		sheet['G3'] = w_avg


		for i in df.index:
			sheet.cell(row=i+3,column=1).value=df['T'][i]
			sheet.cell(row=i+3,column=2).value=df['U'][i]
			sheet.cell(row=i+3,column=3).value=df['V'][i]
			sheet.cell(row=i+3,column=4).value=df['W'][i]
			sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
			sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
			sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
			sheet.cell(row=i+3,column=11).value=find_oct(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
			octant.append(find_oct(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
		
		def name_oct_range(mod=5000):
			octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

			# creating dictionary for mapping 
			MapDic={}                            

			# Creating dictionary with opposite key value pair than 'MapDic'                              
			opp_MapDic={}                                                      
			
			for i in range(4):                                            
				MapDic[i+1]=2*i+1-1                                           
				MapDic[-(i+1)]=2*(i+1)-1
				opp_MapDic[2*i+1-1]=i+1
				opp_MapDic[2*(i+1)-1]=-(i+1)

			# Function to find the rank list from count values of all octs	
			def rankListFind(head_comp):                                     
				temp_head_comp=head_comp.copy()
				temp_head_comp.sort(reverse=True)
				res=[]

				for i in head_comp:
					for j in range(8):
						if(i==temp_head_comp[j]):
							res.append(j+1)
							break
				# Returning the ranked list		
				return res                                                  
			
			# Finding the octant which has rank 1 in the given rank list
			def find_1st_rank(head_comp):                                         
				for i in range(8):
					if(head_comp[i]==1):
						return opp_MapDic[i]

			# Finding the count of rank 1 in the rank 1 mod values of octant x
			def Rank1Count(head_comp,x):                                         
				add=0
				for i in head_comp:
					if(x==i):
						add+=1

				# Return the count		
				return add                                                  
			
			  # Matrix to store rank list for different mod values
			mtrixRank=[] 

			# List to store the octs which have rank 1 in different mod ranges and overall                                               
			rank1_list=[] 
			try:                                                  
				sheet=wb.active
			except:
				print("sheet activation error")	

			# Putting the string 'User Input' at its specified place
			sheet['M4']='Mod '+str(mod)                                           

			# 2-d matrix for storing octs within ranges
			octRangMtrix=[]   

			# Creating a list for storing elements of 9 columns                                                    
			cunt1=[0]*9                                                     

			 # Storing header list in 'cunt1' list
			cunt1[0]='Octant ID'                                           

			for i in range(0,4):
				cunt1[2*i+1]=(i+1)
				cunt1[2*(i+1)]=-(i+1)

			# Appending header list in matrix	
			octRangMtrix.append(cunt1)          

			# Writing header list in worksheet                                  
			for i in range(13,22):                                          
				sheet.cell(row=3,column=i+1).value=cunt1[i-13]
				sheet.cell(row=3,column=i+1).border=border_them
				if(i>13):
					sheet.cell(row=3,column=i+9).value='Rank Octant '+str(cunt1[i-13])
					sheet.cell(row=3,column=i+9).border=border_them
			sheet.cell(row=3,column=31).value='Rank1 Octant ID'
			sheet.cell(row=3,column=32).value='Rank1 Octant Name'
			sheet.cell(row=3,column=31).border=border_them
			sheet.cell(row=3,column=32).border=border_them

			# Resetting values in list 'cunt1'
			cunt1=[0]*9                                                     

			# Finding total count of values in different octs
			for i in octant:                                                
				if(i==1):
					cunt1[1]=cunt1[1]+1
				elif(i==-1):
					cunt1[2]=cunt1[2]+1
				elif(i==2):
					cunt1[3]=cunt1[3]+1
				elif(i==-2):
					cunt1[4]=cunt1[4]+1
				elif(i==3):
					cunt1[5]=cunt1[5]+1
				elif(i==-3):
					cunt1[6]=cunt1[6]+1
				elif(i==4):
					cunt1[7]=cunt1[7]+1
				elif(i==-4):
					cunt1[8]=cunt1[8]+1

			CodeYelow = "00FFFF00"

			# Creating overall count row
			cunt1[0]='Overall Count'                                        
			octRangMtrix.append(cunt1)          

			 # Writing overall count in worksheet                                 
			for i in range(13,22):                                         
				sheet.cell(row=4,column=i+1).value=cunt1[i-13]
				sheet.cell(row=4,column=i+1).border=border_them
			# Removing the header from list	
			cunt1.pop(0)                     

			# Find the rank list                                
			rnk=rankListFind(cunt1)             

			 # Finding the rank 1 octant and appending in rank1_list                      
			rank1_list.append(find_1st_rank(rnk))                         
			mtrixRank.append(rnk)            

			# Writing overall count in worksheet                           
			for i in range(8):                                              
				sheet.cell(row=4,column=23+i).value=mtrixRank[0][i]
				sheet.cell(row=4,column=23+i).border=border_them
				if(mtrixRank[0][i]==1):
					sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
			sheet.cell(row=4,column=31).value=rank1_list[0]
			sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
			sheet.cell(row=4,column=31).border=border_them
			sheet.cell(row=4,column=32).border=border_them
									
			
			# Finding the number of points given in the input
			m=len(octant)     

			# Resetting the values in the list 'count'                                              
			cunt1=[0]*9          

			# Variable to keep track of the index of data we are on                                           
			k=0                     

			# Variable to keep track of row in worksheet                                        
			j=4               

			 # Counting number of values in different octs in mod range                                              
			for i in octant:                                               
				if(i==1):
					cunt1[1]=cunt1[1]+1
				elif(i==-1):
					cunt1[2]=cunt1[2]+1
				elif(i==2):
					cunt1[3]=cunt1[3]+1
				elif(i==-2):
					cunt1[4]=cunt1[4]+1
				elif(i==3):
					cunt1[5]=cunt1[5]+1
				elif(i==-3):
					cunt1[6]=cunt1[6]+1
				elif(i==4):
					cunt1[7]=cunt1[7]+1
				elif(i==-4):
					cunt1[8]=cunt1[8]+1
				k=k+1                                                     
				if(k%mod==1):                                              
					cunt1[0]=str(k-1)+'-'                       
				elif(k%mod==0 or k==m):
					cunt1[0]=cunt1[0]+str(k-1)    

					# Writing the columns of rank, rank1 and octant_name in the worksheet
					for i in range(13,22):                                
						sheet.cell(row=j+1,column=i+1).value=cunt1[i-13]
						sheet.cell(row=j+1,column=i+1).border=border_them
					cunt1.pop(0)                                         
					rnk=rankListFind(cunt1)                           
					rank1_list.append(find_1st_rank(rnk))                
					mtrixRank.append(rnk)                            
					
					# Writing the columns of rank, rank1 and octant_name in the worksheet
					for i in range(8):                                              
						sheet.cell(row=j+1,column=23+i).value=mtrixRank[j-3][i]
						sheet.cell(row=j+1,column=23+i).border=border_them
						if(mtrixRank[j-3][i]==1):
							sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					
					sheet.cell(row=j+1,column=31).value=rank1_list[j-3]
					sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(rank1_list[j-3])]
					sheet.cell(row=j+1,column=31).border=border_them
					sheet.cell(row=j+1,column=32).border=border_them
					
					j=j+1
					octRangMtrix.append(cunt1)
					cunt1=[0]*9                                
													
			rank1_list.pop(0)                                         
			j+=4
			sheet.cell(row=j,column=29).value='Octant ID'        
			sheet.cell(row=j,column=30).value='Octant Name'
			sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
			sheet.cell(row=j,column=29).border=border_them                 
			sheet.cell(row=j,column=30).border=border_them
			sheet.cell(row=j,column=31).border=border_them
			
			# Writing the table of count of rank1 mod values
			for i in range(8):                              
				sheet.cell(row=j+1+i,column=29).value=opp_MapDic[i]
				sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_MapDic[i])]
				sheet.cell(row=j+1+i,column=31).value=Rank1Count(rank1_list,opp_MapDic[i])
				sheet.cell(row=j+1+i,column=29).border=border_them
				sheet.cell(row=j+1+i,column=30).border=border_them
				sheet.cell(row=j+1+i,column=31).border=border_them
		
		def oct_long_subseq_count_with_rang():
			
			# Header list
			r=['Count','Longest Subsequence Length','Count']

			# Writing header of table to worksheet        
			for i in range(3):                                     
				sheet.cell(row=3,column=45+i).value=r[i] 
				sheet.cell(row=3,column=45+i).border=border_them    


			octs=[]

			# Writing octants on leftmost column of the table
			for i in range(2,10,2):                                    
				sheet.cell(row=i+2,column=45).value=i//2
				octs.append(i//2)
				sheet.cell(row=i+3,column=45).value=-(i//2) 
				octs.append(-i//2)
				sheet.cell(row=i+2,column=45).border=border_them
				sheet.cell(row=i+3,column=45).border=border_them                                
			
			# creating dictionary for mapping 
			MapDic={}                                              
			for i in range(0,4):                                            
				MapDic[i+1]=2*i+1-1
				MapDic[-(i+1)]=2*(i+1)-1

			# List for storing number of long subseq		
			cunt1=[0]*8                                             
			longLength=[0]*8                                    
			back=octant[0]

			# Length of curr oct
			l=1                                                           
			n=len(octant)

			 # Temporary variable to store range
			tmpMax=[0]                                                  
			ranges= [[] for x in repeat(None, 8)]            

			 # Loop for finding number and length of long subseq
			for i in range(1,n+1):  
				  # IF last is reached process the whole                
				if(i==n):                                        
					if(longLength[MapDic[back]]<l):                        
						longLength[MapDic[back]]=l
						cunt1[MapDic[back]]=1

						# Writing ending range in temp
						tmpMax.append(df['T'][i-1])            
						ranges[MapDic[back]].clear()         
						ranges[MapDic[back]].append(tmpMax)           
					elif(longLength[MapDic[back]]==l):
						cunt1[MapDic[back]]+=1
						tmpMax.append(df['T'][i-1])
						ranges[MapDic[back]].append(tmpMax) 
				 # If prev and current values are same, increase current length by 1		               
				elif(back==octant[i]):                                      
					l+=1

				# Else process the previous octant values and start with new octant
				else:                                                      
					if(longLength[MapDic[back]]<l):
						longLength[MapDic[back]]=l
						cunt1[MapDic[back]]=1
						ranges[MapDic[back]].clear()               
						tmpMax.append(df['T'][i-1])        
						ranges[MapDic[back]].append(tmpMax)                 
					elif(longLength[MapDic[back]]==l):
						cunt1[MapDic[back]]+=1
						tmpMax.append(df['T'][i-1])
						ranges[MapDic[back]].append(tmpMax)             
					tmpMax=[df['T'][i]]                     
					l=1
					back=octant[i]                            

			# Writing the number and length of longest subsequence in table
			for i in range(2,10):                                    
				sheet.cell(row=i+2,column=46).value=longLength[i-2]
				sheet.cell(row=i+2,column=47).value=cunt1[i-2]
				sheet.cell(row=i+2,column=46).border=border_them
				sheet.cell(row=i+2,column=47).border=border_them
			k=2                                                        
			sheet.cell(row=k+1,column=49).value='Octant ###'                  
			sheet.cell(row=k+1,column=50).value='Longest Subsequence Length'
			sheet.cell(row=k+1,column=51).value='Count'
			sheet.cell(row=k+1,column=49).border=border_them
			sheet.cell(row=k+1,column=50).border=border_them
			sheet.cell(row=k+1,column=51).border=border_them
			
			k+=2
			for i in range(8):
				sheet.cell(row=k,column=49).value=octs[i]            
				sheet.cell(row=k,column=50).value=longLength[i]
				sheet.cell(row=k,column=51).value=cunt1[i]
				sheet.cell(row=k+1,column=49).value='Time'             
				sheet.cell(row=k+1,column=50).value='From'
				sheet.cell(row=k+1,column=51).value='To'
				sheet.cell(row=k,column=49).border=border_them      
				sheet.cell(row=k,column=50).border=border_them
				sheet.cell(row=k,column=51).border=border_them
				sheet.cell(row=k+1,column=49).border=border_them
				sheet.cell(row=k+1,column=50).border=border_them
				sheet.cell(row=k+1,column=51).border=border_them
				x=ranges[i]
				k+=2
				for j in x:
					# Writing ranges in worksheet
					sheet.cell(row=k,column=50).value=j[0]     
					sheet.cell(row=k,column=51).value=j[1]
					sheet.cell(row=k,column=49).border=border_them  
					sheet.cell(row=k,column=50).border=border_them
					sheet.cell(row=k,column=51).border=border_them
					k+=1
			
		def octant_transition_count(mod=5000):
			j=1
			n=len(octant)
			sheet.cell(row=j,column=35).value='Overall Transition Count' 
			sheet.cell(row=j+3,column=34).value='From'
			sheet.cell(row=j+1,column=36).value='To'
			j+=2
			
			octRangMtrix = [ [0]*9 for i in range(9)]                    
			
			# Storing header row and header column in the matrix
			for i in range(0,4):                                        
				octRangMtrix[0][2*i+1]=(i+1)
				octRangMtrix[0][2*(i+1)]=-(i+1)
			for i in range(0,9):
				octRangMtrix[i][0]=octRangMtrix[0][i]
			octRangMtrix[0][0]='Octant #'

			 # creating dictionary for mapping 
			MapDic={}                                             
			for i in range(0,4):
				MapDic[i+1]=2*i+1
				MapDic[-(i+1)]=2*(i+1)

			  # Finding row and column of matrix from transition values
			def find_row_col(x,y):                     
				head_comp=[MapDic[x],MapDic[y]]
				return head_comp
			
			def find_max_ele(head_comp):
				tmpMax=head_comp.copy()
				tmpMax.pop(0)
				large=0
				for i in tmpMax:
					if(large<i):
						large=i
				return large

			back=octant[0]

			# Filling overall transition matrix                                              
			for i in range(1,n):                       
				head_comp=find_row_col(back,octant[i])                        
				octRangMtrix[head_comp[0]][head_comp[1]]+=1
				back=octant[i]

			CodeYelow = "00FFFF00"

			# Writing the overall transition matrix in worksheet
			for i in range(9):                                  
				temp_head_comp=octRangMtrix[i]
				large=find_max_ele(temp_head_comp)
				for k in range(13,22):
					sheet.cell(row=j+i,column=k+22).value=octRangMtrix[i][k-13]
					sheet.cell(row=j+i,column=k+22).border=border_them
					if(i>0 and octRangMtrix[i][k-13]==large):
						sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					if(i!=0 and k!=13):
						octRangMtrix[i][k-13]=0

			# temp-> No. of mod transition tables	
			tmpMax=n//mod+1                                              
			j+=1

			# One iteration for each mod transition table
			for t in range(tmpMax):                                   
				j+=11
				nam1=''

				# Writing Table name in worksheet
				sheet.cell(row=j,column=35).value='Mod Transition Count'     
				sheet.cell(row=j+3,column=34).value='From'
				sheet.cell(row=j+1,column=36).value='To'
				nam1=str(t*mod)+'-'
				if((t+1)*mod-1>n-1):
					nam1+=str(n-1)
				else:
					nam1+=str((t+1)*mod-1)   
				sheet.cell(row=j+1,column=35).value=nam1
				j+=2

				# Incrementing matrix cell corresponding to transition values
				for i in range(t*mod,min(n-1,(t+1)*mod)):                 
					head_comp=find_row_col(octant[i],octant[i+1])
					octRangMtrix[head_comp[0]][head_comp[1]]+=1

				# Writing the transition mod matrix in worksheet
				for i in range(0,9):                                     
					temp_head_comp=octRangMtrix[i]
					if(i>0):
						large=find_max_ele(temp_head_comp)
					for k in range(13,22):
						sheet.cell(row=j+i,column=k+22).value=octRangMtrix[i][k-13]
						sheet.cell(row=j+i,column=k+22).border=border_them
						if(i>0 and octRangMtrix[i][k-13]==large):
							sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						if(i!=0 and k!=13):

							# Resetting matrix for next mod iteration
							octRangMtrix[i][k-13]=0                           


		octant_transition_count(mod)
		name_oct_range(mod)
		oct_long_subseq_count_with_rang()
		s=s.name[:-5]
		path_out='output/'+s+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		wb.save(path_out)

		data_1 = BytesIO()
		wb.save(data_1)
		filename = s+ "cm_vel_octant_analysis_mod_"+str(val)+".xlsx" 
		st.text(filename) 
		st.download_button(label="Download File", file_name=s+
                           "cm_vel_octant_analysis_mod_"+str(val)+".xlsx", data=data_1)
	# a = pathlib.Path(_file_).parent.parent.resolve()
	# input_files = os.path.join(a, "proj2\input")
	# # input_files=os.listdir('input')
	# for i in range(len(input_files)):
	# 	workingFile(mod,file)
	workingFile(file,mod)

#****MULTIPLE file*******

def octant_analysis_mutiple(mod,path):
	from pandas import read_excel
	import openpyxl
	from openpyxl import workbook,load_workbook
	from openpyxl import Workbook
	from openpyxl.styles import PatternFill
	from itertools import repeat

	def workingFile(s,file,mod):							
		inputFilePath=s									
		df=read_excel(inputFilePath)									
		wb=Workbook()											
		sheet=wb.active													
		
		for i in range(11):										
			sheet.cell(row=2,column=i+1).value=head_comp[i]	
		octant=[]
		# sheet.cell(2,1).value = 'T'
		# sheet.cell(2,2).value = 'U'
		# sheet.cell(2,3).value = 'V'
		# sheet.cell(2,4).value = 'W'
		# sheet.cell(2,5).value = 'U Avg'
		# sheet.cell(2,6).value = 'V Avg'
		# sheet.cell(2,7).value = 'W Avg'
		# sheet.cell(2,8).value = "U'=U-U avg"
		# sheet.cell(2,9).value = "V'=V-V avg"
		# sheet.cell(2,10).value = "W'=W-W avg"
		# sheet.cell(2,11).value = 'Octant'

		u_avg=df['U'].mean()                                                 
		v_avg=df['V'].mean()
		w_avg=df['W'].mean()
		sheet.cell(row=1,column=14).value='Overall Octant Count'
		sheet.cell(row=1,column=45).value='Longest Subsequence Length'
		sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
		sheet['E3'] = u_avg
		sheet['F3'] = v_avg
		sheet['G3'] = w_avg


		for i in df.index:
			sheet.cell(row=i+3,column=1).value=df['T'][i]
			sheet.cell(row=i+3,column=2).value=df['U'][i]
			sheet.cell(row=i+3,column=3).value=df['V'][i]
			sheet.cell(row=i+3,column=4).value=df['W'][i]
			sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
			sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
			sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
			sheet.cell(row=i+3,column=11).value=find_oct(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
			octant.append(find_oct(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
		
		def name_oct_range(mod=5000):
			octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

			# creating dictionary for mapping 
			MapDic={}                            

			# Creating dictionary with opposite key value pair than 'MapDic'                              
			opp_MapDic={}                                                      
			
			for i in range(4):                                            
				MapDic[i+1]=2*i+1-1                                           
				MapDic[-(i+1)]=2*(i+1)-1
				opp_MapDic[2*i+1-1]=i+1
				opp_MapDic[2*(i+1)-1]=-(i+1)

			# Function to find the rank list from count values of all octs	
			def rankListFind(head_comp):                                     
				temp_head_comp=head_comp.copy()
				temp_head_comp.sort(reverse=True)
				res=[]

				for i in head_comp:
					for j in range(8):
						if(i==temp_head_comp[j]):
							res.append(j+1)
							break
				# Returning the ranked list		
				return res                                                  
			
			# Finding the octant which has rank 1 in the given rank list
			def find_1st_rank(head_comp):                                         
				for i in range(8):
					if(head_comp[i]==1):
						return opp_MapDic[i]

			# Finding the count of rank 1 in the rank 1 mod values of octant x
			def Rank1Count(head_comp,x):                                         
				add=0
				for i in head_comp:
					if(x==i):
						add+=1

				# Return the count		
				return add                                                  
			
			  # Matrix to store rank list for different mod values
			mtrixRank=[] 

			# List to store the octs which have rank 1 in different mod ranges and overall                                               
			rank1_list=[] 
			try:                                                  
				sheet=wb.active
			except:
				print("sheet activation error")	

			# Putting the string 'User Input' at its specified place
			sheet['M4']='Mod '+str(mod)                                           

			# 2-d matrix for storing octs within ranges
			octRangMtrix=[]   

			# Creating a list for storing elements of 9 columns                                                    
			cunt1=[0]*9                                                     

			 # Storing header list in 'cunt1' list
			cunt1[0]='Octant ID'                                           

			for i in range(0,4):
				cunt1[2*i+1]=(i+1)
				cunt1[2*(i+1)]=-(i+1)

			# Appending header list in matrix	
			octRangMtrix.append(cunt1)          

			# Writing header list in worksheet                                  
			for i in range(13,22):                                          
				sheet.cell(row=3,column=i+1).value=cunt1[i-13]
				sheet.cell(row=3,column=i+1).border=border_them
				if(i>13):
					sheet.cell(row=3,column=i+9).value='Rank Octant '+str(cunt1[i-13])
					sheet.cell(row=3,column=i+9).border=border_them
			sheet.cell(row=3,column=31).value='Rank1 Octant ID'
			sheet.cell(row=3,column=32).value='Rank1 Octant Name'
			sheet.cell(row=3,column=31).border=border_them
			sheet.cell(row=3,column=32).border=border_them

			# Resetting values in list 'cunt1'
			cunt1=[0]*9                                                     

			# Finding total count of values in different octs
			for i in octant:                                                
				if(i==1):
					cunt1[1]=cunt1[1]+1
				elif(i==-1):
					cunt1[2]=cunt1[2]+1
				elif(i==2):
					cunt1[3]=cunt1[3]+1
				elif(i==-2):
					cunt1[4]=cunt1[4]+1
				elif(i==3):
					cunt1[5]=cunt1[5]+1
				elif(i==-3):
					cunt1[6]=cunt1[6]+1
				elif(i==4):
					cunt1[7]=cunt1[7]+1
				elif(i==-4):
					cunt1[8]=cunt1[8]+1

			CodeYelow = "00FFFF00"

			# Creating overall count row
			cunt1[0]='Overall Count'                                        
			octRangMtrix.append(cunt1)          

			 # Writing overall count in worksheet                                 
			for i in range(13,22):                                         
				sheet.cell(row=4,column=i+1).value=cunt1[i-13]
				sheet.cell(row=4,column=i+1).border=border_them
			# Removing the header from list	
			cunt1.pop(0)                     

			# Find the rank list                                
			rnk=rankListFind(cunt1)             

			 # Finding the rank 1 octant and appending in rank1_list                      
			rank1_list.append(find_1st_rank(rnk))                         
			mtrixRank.append(rnk)            

			# Writing overall count in worksheet                           
			for i in range(8):                                              
				sheet.cell(row=4,column=23+i).value=mtrixRank[0][i]
				sheet.cell(row=4,column=23+i).border=border_them
				if(mtrixRank[0][i]==1):
					sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
			sheet.cell(row=4,column=31).value=rank1_list[0]
			sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
			sheet.cell(row=4,column=31).border=border_them
			sheet.cell(row=4,column=32).border=border_them
									
			
			# Finding the number of points given in the input
			m=len(octant)     

			# Resetting the values in the list 'count'                                              
			cunt1=[0]*9          

			# Variable to keep track of the index of data we are on                                           
			k=0                     

			# Variable to keep track of row in worksheet                                        
			j=4               

			 # Counting number of values in different octs in mod range                                              
			for i in octant:                                               
				if(i==1):
					cunt1[1]=cunt1[1]+1
				elif(i==-1):
					cunt1[2]=cunt1[2]+1
				elif(i==2):
					cunt1[3]=cunt1[3]+1
				elif(i==-2):
					cunt1[4]=cunt1[4]+1
				elif(i==3):
					cunt1[5]=cunt1[5]+1
				elif(i==-3):
					cunt1[6]=cunt1[6]+1
				elif(i==4):
					cunt1[7]=cunt1[7]+1
				elif(i==-4):
					cunt1[8]=cunt1[8]+1
				k=k+1                                                     
				if(k%mod==1):                                              
					cunt1[0]=str(k-1)+'-'                       
				elif(k%mod==0 or k==m):
					cunt1[0]=cunt1[0]+str(k-1)    

					# Writing the columns of rank, rank1 and octant_name in the worksheet
					for i in range(13,22):                                
						sheet.cell(row=j+1,column=i+1).value=cunt1[i-13]
						sheet.cell(row=j+1,column=i+1).border=border_them
					cunt1.pop(0)                                         
					rnk=rankListFind(cunt1)                           
					rank1_list.append(find_1st_rank(rnk))                
					mtrixRank.append(rnk)                            
					
					# Writing the columns of rank, rank1 and octant_name in the worksheet
					for i in range(8):                                              
						sheet.cell(row=j+1,column=23+i).value=mtrixRank[j-3][i]
						sheet.cell(row=j+1,column=23+i).border=border_them
						if(mtrixRank[j-3][i]==1):
							sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					
					sheet.cell(row=j+1,column=31).value=rank1_list[j-3]
					sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(rank1_list[j-3])]
					sheet.cell(row=j+1,column=31).border=border_them
					sheet.cell(row=j+1,column=32).border=border_them
					
					j=j+1
					octRangMtrix.append(cunt1)
					cunt1=[0]*9                                
													
			rank1_list.pop(0)                                         
			j+=4
			sheet.cell(row=j,column=29).value='Octant ID'        
			sheet.cell(row=j,column=30).value='Octant Name'
			sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
			sheet.cell(row=j,column=29).border=border_them                 
			sheet.cell(row=j,column=30).border=border_them
			sheet.cell(row=j,column=31).border=border_them
			
			# Writing the table of count of rank1 mod values
			for i in range(8):                              
				sheet.cell(row=j+1+i,column=29).value=opp_MapDic[i]
				sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_MapDic[i])]
				sheet.cell(row=j+1+i,column=31).value=Rank1Count(rank1_list,opp_MapDic[i])
				sheet.cell(row=j+1+i,column=29).border=border_them
				sheet.cell(row=j+1+i,column=30).border=border_them
				sheet.cell(row=j+1+i,column=31).border=border_them
		
		def oct_long_subseq_count_with_rang():
			
			# Header list
			r=['Count','Longest Subsequence Length','Count']

			# Writing header of table to worksheet        
			for i in range(3):                                     
				sheet.cell(row=3,column=45+i).value=r[i] 
				sheet.cell(row=3,column=45+i).border=border_them    


			octs=[]

			# Writing octants on leftmost column of the table
			for i in range(2,10,2):                                    
				sheet.cell(row=i+2,column=45).value=i//2
				octs.append(i//2)
				sheet.cell(row=i+3,column=45).value=-(i//2) 
				octs.append(-i//2)
				sheet.cell(row=i+2,column=45).border=border_them
				sheet.cell(row=i+3,column=45).border=border_them                                
			
			# creating dictionary for mapping 
			MapDic={}                                              
			for i in range(0,4):                                            
				MapDic[i+1]=2*i+1-1
				MapDic[-(i+1)]=2*(i+1)-1

			# List for storing number of long subseq		
			cunt1=[0]*8                                             
			longLength=[0]*8                                    
			back=octant[0]

			# Length of curr oct
			l=1                                                           
			n=len(octant)

			 # Temporary variable to store range
			tmpMax=[0]                                                  
			ranges= [[] for x in repeat(None, 8)]            

			 # Loop for finding number and length of long subseq
			for i in range(1,n+1):  
				  # IF last is reached process the whole                
				if(i==n):                                        
					if(longLength[MapDic[back]]<l):                        
						longLength[MapDic[back]]=l
						cunt1[MapDic[back]]=1

						# Writing ending range in temp
						tmpMax.append(df['T'][i-1])            
						ranges[MapDic[back]].clear()         
						ranges[MapDic[back]].append(tmpMax)           
					elif(longLength[MapDic[back]]==l):
						cunt1[MapDic[back]]+=1
						tmpMax.append(df['T'][i-1])
						ranges[MapDic[back]].append(tmpMax) 
				 # If prev and current values are same, increase current length by 1		               
				elif(back==octant[i]):                                      
					l+=1

				# Else process the previous octant values and start with new octant
				else:                                                      
					if(longLength[MapDic[back]]<l):
						longLength[MapDic[back]]=l
						cunt1[MapDic[back]]=1
						ranges[MapDic[back]].clear()               
						tmpMax.append(df['T'][i-1])        
						ranges[MapDic[back]].append(tmpMax)                 
					elif(longLength[MapDic[back]]==l):
						cunt1[MapDic[back]]+=1
						tmpMax.append(df['T'][i-1])
						ranges[MapDic[back]].append(tmpMax)             
					tmpMax=[df['T'][i]]                     
					l=1
					back=octant[i]                            

			# Writing the number and length of longest subsequence in table
			for i in range(2,10):                                    
				sheet.cell(row=i+2,column=46).value=longLength[i-2]
				sheet.cell(row=i+2,column=47).value=cunt1[i-2]
				sheet.cell(row=i+2,column=46).border=border_them
				sheet.cell(row=i+2,column=47).border=border_them
			k=2                                                        
			sheet.cell(row=k+1,column=49).value='Octant ###'                  
			sheet.cell(row=k+1,column=50).value='Longest Subsequence Length'
			sheet.cell(row=k+1,column=51).value='Count'
			sheet.cell(row=k+1,column=49).border=border_them
			sheet.cell(row=k+1,column=50).border=border_them
			sheet.cell(row=k+1,column=51).border=border_them
			
			k+=2
			for i in range(8):
				sheet.cell(row=k,column=49).value=octs[i]            
				sheet.cell(row=k,column=50).value=longLength[i]
				sheet.cell(row=k,column=51).value=cunt1[i]
				sheet.cell(row=k+1,column=49).value='Time'             
				sheet.cell(row=k+1,column=50).value='From'
				sheet.cell(row=k+1,column=51).value='To'
				sheet.cell(row=k,column=49).border=border_them      
				sheet.cell(row=k,column=50).border=border_them
				sheet.cell(row=k,column=51).border=border_them
				sheet.cell(row=k+1,column=49).border=border_them
				sheet.cell(row=k+1,column=50).border=border_them
				sheet.cell(row=k+1,column=51).border=border_them
				x=ranges[i]
				k+=2
				for j in x:
					# Writing ranges in worksheet
					sheet.cell(row=k,column=50).value=j[0]     
					sheet.cell(row=k,column=51).value=j[1]
					sheet.cell(row=k,column=49).border=border_them  
					sheet.cell(row=k,column=50).border=border_them
					sheet.cell(row=k,column=51).border=border_them
					k+=1
			
		def octant_transition_count(mod=5000):
			j=1
			n=len(octant)
			sheet.cell(row=j,column=35).value='Overall Transition Count' 
			sheet.cell(row=j+3,column=34).value='From'
			sheet.cell(row=j+1,column=36).value='To'
			j+=2
			
			octRangMtrix = [ [0]*9 for i in range(9)]                    
			
			# Storing header row and header column in the matrix
			for i in range(0,4):                                        
				octRangMtrix[0][2*i+1]=(i+1)
				octRangMtrix[0][2*(i+1)]=-(i+1)
			for i in range(0,9):
				octRangMtrix[i][0]=octRangMtrix[0][i]
			octRangMtrix[0][0]='Octant #'

			 # creating dictionary for mapping 
			MapDic={}                                             
			for i in range(0,4):
				MapDic[i+1]=2*i+1
				MapDic[-(i+1)]=2*(i+1)

			  # Finding row and column of matrix from transition values
			def find_row_col(x,y):                     
				head_comp=[MapDic[x],MapDic[y]]
				return head_comp
			
			def find_max_ele(head_comp):
				tmpMax=head_comp.copy()
				tmpMax.pop(0)
				large=0
				for i in tmpMax:
					if(large<i):
						large=i
				return large

			back=octant[0]

			# Filling overall transition matrix                                              
			for i in range(1,n):                       
				head_comp=find_row_col(back,octant[i])                        
				octRangMtrix[head_comp[0]][head_comp[1]]+=1
				back=octant[i]

			CodeYelow = "00FFFF00"

			# Writing the overall transition matrix in worksheet
			for i in range(9):                                  
				temp_head_comp=octRangMtrix[i]
				large=find_max_ele(temp_head_comp)
				for k in range(13,22):
					sheet.cell(row=j+i,column=k+22).value=octRangMtrix[i][k-13]
					sheet.cell(row=j+i,column=k+22).border=border_them
					if(i>0 and octRangMtrix[i][k-13]==large):
						sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					if(i!=0 and k!=13):
						octRangMtrix[i][k-13]=0

			# temp-> No. of mod transition tables	
			tmpMax=n//mod+1                                              
			j+=1

			# One iteration for each mod transition table
			for t in range(tmpMax):                                   
				j+=11
				nam1=''

				# Writing Table name in worksheet
				sheet.cell(row=j,column=35).value='Mod Transition Count'     
				sheet.cell(row=j+3,column=34).value='From'
				sheet.cell(row=j+1,column=36).value='To'
				nam1=str(t*mod)+'-'
				if((t+1)*mod-1>n-1):
					nam1+=str(n-1)
				else:
					nam1+=str((t+1)*mod-1)   
				sheet.cell(row=j+1,column=35).value=nam1
				j+=2

				# Incrementing matrix cell corresponding to transition values
				for i in range(t*mod,min(n-1,(t+1)*mod)):                 
					head_comp=find_row_col(octant[i],octant[i+1])
					octRangMtrix[head_comp[0]][head_comp[1]]+=1

				# Writing the transition mod matrix in worksheet
				for i in range(0,9):                                     
					temp_head_comp=octRangMtrix[i]
					if(i>0):
						large=find_max_ele(temp_head_comp)
					for k in range(13,22):
						sheet.cell(row=j+i,column=k+22).value=octRangMtrix[i][k-13]
						sheet.cell(row=j+i,column=k+22).border=border_them
						if(i>0 and octRangMtrix[i][k-13]==large):
							sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						if(i!=0 and k!=13):

							# Resetting matrix for next mod iteration
							octRangMtrix[i][k-13]=0                           


		octant_transition_count(mod)
		name_oct_range(mod)
		oct_long_subseq_count_with_rang()
		file=file[:-5]
		path_out='output/'+file+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		wb.save(path_out)
		data_1 = BytesIO()
		wb.save(data_1)
		filename = file +'cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		st.text(filename)
		st.download_button(label="Download File", file_name=filename   , data=data_1)
	for file in os.listdir(path):
		a = path+"\\"+file
		workingFile(a,file,mod)

st.header("Project - 2")
val = st.number_input(label="Enter Mod Value", min_value=1, step=1)
file = st.file_uploader("Please choose a file", type=["xlsx"])
if os.path.exists("output"):
    pass

else:
	a = pathlib.Path(__file__).parent.parent.resolve()
	b = os.path.join(a,"proj2") 
	filename = os.path.join(b,"output") 
	os.mkdir(filename)

if st.button("Compute"):
	a = pathlib.Path(__file__).parent.parent.resolve()
	d = os.path.join(a,"proj2\output")
	for f in os.listdir(d):
		os.remove(os.path.join(d, f))
	octant_analysis_single(val,file)


direc = st.text_input(label="path of directiory")
if st.button("Compute_1"):
	octant_analysis_mutiple(val,direc)