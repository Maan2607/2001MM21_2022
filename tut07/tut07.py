from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.borders import Border, Side
start_time = datetime.now()

# Help


def octant_analysis(mod=5000):
	from openpyxl import load_workbook
	import os
	import numpy as np
	from operator import itemgetter
	
	os.system('cls')
	curr = os.getcwd()  # copy current folder adderss in curr
	if os.path.exists("output"):
		for f in os.listdir("output"):# delete all file in output folder
			os.remove(os.path.join("output", f))
		os.rmdir("output")  # delete output folder
	os.mkdir(curr.replace('\\', '/')+"/output/")  # make output folder
	os.chdir("input")
	fil = os. listdir()  # storing all file name of input folder in list
	os.chdir(curr)
	for list in fil:
		print(list)
		os.chdir("input")  # taking input
		wb = load_workbook(list)
		sheet = wb.active
		time1 = []  # make a new list
		v1 = []    # make a new list
		u1 = []    # make a new list
		w1 = []    # make a new list
		v2 = []    # make a list for difference
		u2 = []    # make a list for difference
		w2 = []    # make a list for difference
		oct = []    # make a list for octant stori
		wb = load_workbook(list)

		sheet = wb.active
		n = sheet.max_row  # n is the number of rows of data
		for r in range(2, n+1):       # making of time1 list from given data
			time1.append(sheet.cell(row=r, column=1).value)
			u1.append(sheet.cell(row=r, column=2).value)
			v1.append(sheet.cell(row=r, column=3).value)# making of w1 list from given data
			w1.append(sheet.cell(row=r, column=4).value)
		v_mean = np.mean(v1, dtype=np.float64)  # calculate the mean of v1
		u_mean = np.mean(u1, dtype=np.float64)  # mean of u1
		w_mean = np.mean(w1, dtype=np.float64)  # mean of w1
		for r in range(2, n+1):
			u2.append(float(sheet.cell(row=r, column=2).value) -u_mean)  # pushing the differences
			v2.append(float(sheet.cell(row=r, column=3).value) -v_mean)  # pushing the differences
			w2.append(float(sheet.cell(row=r, column=4).value) -w_mean)  # pushing the differenc
		tu3=[]
		for i in range(0, n-1):  # using this for loop we will get octant
			if ((u2[i] >= 0) and (v2[i] >= 0) and (w2[i] >= 0)):
				oct.append(1)
			elif ((u2[i] >= 0) and (v2[i] >= 0) and (w2[i] < 0)):
				oct.append(-1)
			elif ((u2[i] < 0) and (v2[i] >= 0) and (w2[i] >= 0)):
				oct.append(2)
			elif ((u2[i] < 0) and (v2[i] >= 0) and (w2[i] < 0)):
				oct.append(-2)
			elif ((u2[i] >= 0) and (v2[i] < 0) and (w2[i] >= 0)):
				oct.append(4)
			elif ((u2[i] >= 0) and (v2[i] < 0) and (w2[i] < 0)):
				oct.append(-4)
			elif ((u2[i] < 0) and (v2[i] < 0) and (w2[i] >= 0)):
				oct.append(3)
			else:
				oct.append(-3)
		s1 = 0  # it is used to store the  longest subsequence value of octant 1
		s2 = 0  # it is used to store the longest subsequence value of octant -1
		s3 = 0  # it is used to store the longest subsequence value of octant 2
		s4 = 0  # it is used to store the longest subsequence value of octant -2
		s5 = 0  # it is used to store the longest subsequence value of octant 3
		s6 = 0  # it is used to store the longest subsequence value of octant -3
		s7 = 0  # it is used to store the longest subsequence value of octant 4
		s8 = 0  # it is used to store the longest subsequence value of octant -4
		a = 0  # It will give the longest subsequence length for octant 1
		b = 0  # It will give the longest subsequence length for octant -1
		c = 0  # It will give the longest subsequence length for octant 2
		d = 0  # It will give the longest subsequence length for octant -2
		e = 0  # It will give the longest subsequence length for octant 3
		f = 0  # It will give the longest subsequence length for octant -3
		g = 0  # It will give the longest subsequence length for octant 4
		h = 0  # It will give the longest subsequence length for octant -4
		c1 = 0  # It will give the count for octant 1
		c2 = 0  # It will give the count for octant -1
		c3 = 0  # It will give the count for octant 2
		c4 = 0  # It will give the count for octant -2
		c5 = 0  # It will give the count for octant 3
		c6 = 0  # It will give the count for octant -3
		c7 = 0  # It will give the count for octant 4
		c8 = 0  # It will give the count for octant -4 
		for r in range(0, n-1):
			if oct[r] == 1:
				s1 = s1+1
				if r == n-2:
					if s1 > a:
						a = s1
						s1 = 0
						c1 = 1
					elif a > s1:
						s1 = 0
					else:
						s1 = 0
						c1 = c1+1
			else:
				if s1 > a:
					a = s1
					s1 = 0
					c1 = 1
				elif a > s1:
					s1 = 0
				else:
					s1 = 0
					c1 = c1+1
		cn = 0  # initiallize count check with 0
		t1 = []  # This list will store the starting time for longest subsequence length for octant +1
		for r in range(0, n-1):
			if oct[r] == 1:
				if cn == 0:
					ti1 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == a:  # Checking if length is equal to longest subsequence length
				t1.append(ti1)
				cn = 0
			
						# This for loop will solve the count and longest subsequence length for octant -1
		for r in range(0, n-1):
			if oct[r] == -1:
				s2 = s2+1
				if r == n-2:
					if s2 > b:
						b = s2
						s2 = 0
						c2 = 1
					elif b > s2:
						s2 = 0
					else:
						s2 = 0
						c2 = c2+1
			else:
				if s2 > b:
					b = s2
					s2 = 0
					c2 = 1
				elif b > s2:
					s2 = 0
				else:
					s2 = 0
					c2 = c2+1
					
		cn = 0  # initiallize count check with 0
		t2 = []  # This list will store the starting time for longest subsequence length for octant -1
		for r in range(0, n-1):
			if oct[r] == -1:
				if cn == 0:
					ti2 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == b:  # Checking if length is equal to longest subsequence length
				t2.append(ti2)
				cn = 0

    			
					# This for loop will solve the count and longest subsequence length for octant 2
		for r in range(0, n-1):
			if oct[r] == 2:
				s3 = s3+1
				if r == n-2:
					if s3 > c:
						c = s3
						s3 = 0
						c3 = 1
					elif c > s3:
						s3 = 0
					else:
						s3 = 0
						c3 = c3+1
			else:
				if s3 > c:
					c = s3
					s3 = 0
					c3 = 1
				elif c > s3:
					s3 = 0
				else:
					s3 = 0
					c3 = c3+1
		cn = 0  # initiallize count check with 0
		t3 = []   # This list will store the starting time for longest subsequence length for octant 2
		for r in range(0, n-1):
			if oct[r] == 2:
				if cn == 0:
					ti3 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == c:  # Checking if length is equal to longest subsequence length
				t3.append(ti3)
				cn = 0

		for r in range(0, n-1):
			if oct[r] == -2:
				s4 = s4+1
				if r == n-2:
					if s4 > d:
						d = s4
						s4 = 0
						c4 = 1
					elif d > s4:
						s4 = 0
					else:
						s4 = 0
						c4 = c4+1
			else:
				if s4 > d:
					d = s4
					s4 = 0
					c4 = 1
				elif d > s4:
					s4 = 0
				else:
					s4 = 0
					c4 = c4+1
				# This for loop will solve the count and longest subsequence length for octant 3
		cn = 0  # initiallize count check with 0
		t4 = []  # This list will store the starting time for longest subsequence length for octant -2
		for r in range(0, n-1):
			if oct[r] == -2:
				if cn == 0:
					ti4 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == d:  # Checking if length is equal to longest subsequence length
				t4.append(ti4)
				cn = 0

		
		
		for r in range(0, n-1):
			if oct[r] == 3:
				s5 = s5+1
				if r == n-2:
					if s5 > e:
						e = s5
						s5 = 0
						c5 = 1
					elif e > s5:
						s5 = 0
					else:
						s5 = 0
						c5 = c5+1
			else:
				if s5 > e:
					e = s5
					s5 = 0
					c5 = 1
				elif e > s5:
					s5 = 0
				else:
					s5 = 0
					c5 = c5+1
				# This for loop will solve the count and longest subsequence length for octant -3
		cn = 0  # initiallize count check with 0
		t5 = []  # This list will store the starting time for longest subsequence length for octant 3
		for r in range(0, n-1):
			if oct[r] == 3:
				if cn == 0:
					ti5 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == e:  # Checking if length is equal to longest subsequence length
				t5.append(ti5)
				cn = 0

		
		for r in range(0, n-1):
			if oct[r] == -3:
				s6 = s6+1
				if r == n-2:
					if s6 > a:
						f = s6
						s6 = 0
						c6 = 1
					elif f > s6:
						s6 = 0
					else:
						s6 = 0
						c6 = c6+1
			else:
				if s6 > f:
					f = s6
					s6 = 0
					c6 = 1
				elif f > s6:
					s6 = 0
				else:
					s6 = 0
					c6 = c6+1
			# This for loop will solve the count and longest subsequence length for octant 4
		cn = 0  # initiallize count check with 0
		t6 = []  # This list will store the starting time for longest subsequence length for octant -3
		for r in range(0, n-1):
			if oct[r] == -3:
				if cn == 0:
					ti6 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == f:  # Checking if length is equal to longest subsequence length
				t6.append(ti6)
				cn = 0

		
		
		for r in range(0, n-1):
			if oct[r] == 4:
				s7 = s7+1
				if r == n-2:
					if s7 > g:
						g = s7
						s7 = 0
						c7 = 1
					elif g > s7:
						s7 = 0
					else:
						s7 = 0
						c7 = c7+1
			else:
				if s7 > g:
					g = s7
					s7 = 0
					c7 = 1
				elif g > s7:
					s7 = 0
				else:
					s7 = 0
					c7 = c7+1
		cn = 0  # initiallize count check with 0
		t7 = []  # This list will store the starting time for longest subsequence length for octant 4
		for r in range(0, n-1):
			if oct[r] == 4:
				if cn == 0:
					ti7 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == g:  # Checking if length is equal to longest subsequence length
				t7.append(ti7)
				cn = 0
			
				# This for loop will solve the count and longest subsequence length for octant -4
		for r in range(0, n-1):
			if oct[r] == -4:
				s8 = s8+1
				if r == n-2:
					if s8 > a:
						h = s8
						s8 = 0
						c8 = 1
					elif h > s8:
						s8 = 0
					else:
						s8 = 0
						c8 = c8+1
			else:
				if s8 > h:
					h = s8
					s8 = 0
					c8 = 1
				elif h > s8:
					s8 = 0
				else:
					s8 = 0
					c8 = c8+1
		cn = 0  # initiallize count check with 0
		t8 = []  # This list will store the starting time for longest subsequence length for octant -4
		for r in range(0, n-1):
			if oct[r] == -4:
				if cn == 0:
					ti8 = time1[r]
				cn = cn+1
			else:
				cn = 0
			if cn == h:  # Checking if length is equal to longest subsequence length
				t8.append(ti8)
				cn = 0
			
		tu3.append(["Octant##","Longest Subsequence length","Count"])
		tu3.append(["1",a,c1])
		tu3.append(["-1",b,c2])
		tu3.append(["2",c,c3])
		tu3.append(["-2",d,c4])
		tu3.append(["3",e,c5])
		tu3.append(["-3",f,c6])
		tu3.append(["4",g,c7])
		tu3.append(["-4",h,c8])
		for i in range(0,1000):
			tu3.append(["","",""])
		e1 = []  # This list is made to print the first column the new required table
		e1.append("Octant###")
		e1.append("1")
		e1.append("Time")
		for i in range(0, c1):  # This loop is for giving c1 no. of space in first coloum
			e1.append("")
		e1.append("-1")
		e1.append("Time")
		for i in range(0, c2):  # This loop is for giving c2 no. of space in first coloum
			e1.append("")
		e1.append("2")
		e1.append("Time")
		for i in range(0, c3):  # This loop is for giving c3 no. of space in first coloum
			e1.append("")
		e1.append("-2")
		e1.append("Time")
		for i in range(0, c4):  # This loop is for giving c4 no. of space in first coloum
			e1.append("")
		e1.append("3")
		e1.append("Time")
		for i in range(0, c5):  # This loop is for giving c5 no. of space in first coloum
			e1.append("")
		e1.append("-3")
		e1.append("Time")
		for i in range(0, c6):  # This loop is for giving c6 no. of space in first coloum
			e1.append("")
		e1.append("4")
		e1.append("Time")
		for i in range(0, c7):  # This loop is for giving c7 no. of space in first coloum
			e1.append("")
		e1.append("-4")
		e1.append("Time")
		for i in range(0, c8):  # This loop is for giving c8 no. of space in first coloum
			e1.append("")
		v = len(e1)
        
		b1 = []  # This list is made to print the second coloumn new required table
		b1.append("Longest Subsequence Length")
		b1.append(a)
		b1.append("From")
		for i in range(0, c1):  # This loop is for giving starting time of longest subsequece length for octant +1 in second coloum
			b1.append(t1[i])
		b1.append(b)
		b1.append("From")
		for i in range(0, c2):  # This loop is for giving starting time of longest subsequece length for octant -1 in second coloum
			b1.append(t2[i])
		b1.append(c)
		b1.append("From")
		for i in range(0, c3):  # This loop is for giving starting time of longest subsequece length for octant +2 in second coloum
			b1.append(t3[i])
		b1.append(d)
		b1.append("From")
		for i in range(0, c4):  # This loop is for giving starting time of longest subsequece length for octant -2 in second coloum
			b1.append(t4[i])
		b1.append(e)
		b1.append("From")
		for i in range(0, c5):  # This loop is for giving starting time of longest subsequece length for octant +3 in second coloum
			b1.append(t5[i])
		b1.append(f)
		b1.append("From")
		for i in range(0, c6):  # This loop is for giving starting time of longest subsequece length for octant -3 in second coloum
			b1.append(t6[i])
		b1.append(g)
		b1.append("From")
		for i in range(0, c7):  # This loop is for giving starting time of longest subsequece length for octant +4 in second coloum
			b1.append(t7[i])
		b1.append(h)
		b1.append("From")
		for i in range(0, c8):  # This loop is for giving starting time of longest subsequece length for octant -4 in second coloum
			b1.append(t8[i])
        

		d1 = []  # This list is made to print third coloumn the new required table
		d1.append("Count")
		d1.append(c1)
		d1.append("To")
		for i in range(0, c1):  # This loop is for giving ending time of longest subsequece length for octant +1 in third coloum
			d1.append(t1[i] + 0.01*(a-1))
		d1.append(c2)
		d1.append("To")
		for i in range(0, c2):  # This loop is for giving ending time of longest subsequece length for octant -1 in third coloum
			d1.append(t2[i] + 0.01*(b-1))
		d1.append(c3)
		d1.append("To")
		for i in range(0, c3):  # This loop is for giving ending time of longest subsequece length for octant +2 in third coloum
			d1.append(t3[i] + 0.01*(c-1))
		d1.append(c4)
		d1.append("To")
		for i in range(0, c4):  # This loop is for giving ending time of longest subsequece length for octant -2 in third coloum
			d1.append(t4[i] + 0.01*(d-1))
		d1.append(c5)
		d1.append("To")
		for i in range(0, c5):  # This loop is for giving ending time of longest subsequece length for octant +3 in third coloum
			d1.append(t5[i] + 0.01*(e-1))
		d1.append(c6)
		d1.append("To")
		for i in range(0, c6):  # This loop is for giving ending time of longest subsequece length for octant -3 in third coloum
			d1.append(t6[i] + 0.01*(f-1))
		d1.append(c7)
		d1.append("To")
		for i in range(0, c7):  # This loop is for giving ending time of longest subsequece length for octant +4 in third coloum
			d1.append(t7[i] + 0.01*(g-1))
		d1.append(c8)
		d1.append("To")
		for i in range(0, c8):  # This loop is for giving ending time of longest subsequece length for octant +4 in third coloum
			d1.append(t8[i] + 0.01*(h-1))
        
		for i in range(0,1000):
			e1.append("")
			b1.append("")
			d1.append("")
		c1 = 0  # overall count of octant 1
		c2 = 0  # overall count of octant -1
		c3 = 0  # overall count of octant 2
		c4 = 0  # overall count of octant -2
		c5 = 0  # overall count of octant 3
		c6 = 0  # overall count of octant -3
		c7 = 0  # overall count of octant 4
		c8 = 0  # overall count of octant -4
		for i in range(0, n-1):  # using this for loop we will get octant
			if (oct[i] == 1):
				c1 = c1+1
			elif (oct[i] == -1):
				c2 = c2+1
			elif (oct[i] == 2):
				c3 = c3+1
			elif (oct[i] == -2):
				c4 = c4+1
			elif (oct[i] == 3):
				c5 = c5+1
			elif (oct[i] == -3):
				c6 = c6+1
			elif (oct[i] == 4):
				c7 = c7+1
			else:
				c8 = c8+1
		rank = []  # this list will give the ranks of all octants in differnt intervals
		rank1_id_name = []  # This list will help rank list in storing the data
		req = []  # This list will help rank list in storing the data
		req.append([c1, 1])
		req.append([c2, 2])
		req.append([c3, 3])
		req.append([c4, 4])
		req.append([c5, 5])
		req.append([c6, 6])
		req.append([c7, 7])
		req.append([c8, 8])
		req.sort()
		# this will give ranks of different octant in a particular interval
		req_ranks = [0, 0, 0, 0, 0, 0, 0, 0]
		req_ranks[req[0][1]-1] = 8  # Giving rank 8 in sorted list
		req_ranks[req[1][1]-1] = 7  # Giving rank 7 in sorted list
		req_ranks[req[2][1]-1] = 6  # Giving rank 6 in sorted list
		req_ranks[req[3][1]-1] = 5  # Giving rank 5 in sorted list
		req_ranks[req[4][1]-1] = 4  # Giving rank 4 in sorted list
		req_ranks[req[5][1]-1] = 3  # Giving rank 3 in sorted list
		req_ranks[req[6][1]-1] = 2  # Giving rank 2 in sorted list
		req_ranks[req[7][1]-1] = 1  # Giving rank 1 in sorted list
		rank.append(req_ranks)  # Appending this list in rank list
		rank_id_name = []
		if req_ranks[0] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [1, "Internal outward Interaction"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[1] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [-1, "External outward Interaction"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[2] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [2, "External Ejection"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[3] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [-2, "Internal Ejection"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[4] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [3, "External inward Interaction"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[5] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [-3, "Internal inward Interaction"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[6] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [4, "Internal Sweep"]
			rank1_id_name.append(rank_id_name)
		elif req_ranks[7] == 1:  # Givng name and id to Rank 1 Octant
			rank_id_name = [-4, "External Sweep"]
			rank1_id_name.append(rank_id_name)
		o1=[] #make a list of octant count in mod range 
		o2=[] #make a list of octant count in mod range
		o3=[] #make a list of octant count in mod range
		o4=[] #make a list of octant count in mod range
		o5=[] #make a list of octant count in mod range
		o6=[] #make a list of octant count in mod range
		o7=[] #make a list of octant count in mod range
		o8=[] #make a list of octant count in mod range
		oct1 = 0  # making variables for using in function
		oct2 = 0  # making variables for using in function
		oct3 = 0  # making variables for using in function
		oct4 = 0  # making variables for using in function
		oct5 = 0  # making variables for using in function
		oct6 = 0  # making variables for using in function
		oct7 = 0  # making variables for using in function
		oct8 = 0  # making variables for using in function
		m = int((n-2)/mod) + 1  # m will give no.of interval
		sn1 = 0  # This will give Count of rank 1 if Octant 1 in different mod intervals
		sn2 = 0  # This will give Count of rank -1 if Octant 1 in different mod intervals
		sn3 = 0  # This will give Count of rank 2if Octant 1 in different mod intervals
		sn4 = 0  # This will give Count of rank -2 if Octant 1 in different mod intervals
		sn5 = 0  # This will give Count of rank 3 if Octant 1 in different mod intervals
		sn6 = 0  # This will give Count of rank -3 if Octant 1 in different mod intervals
		sn7 = 0  # This will give Count of rank 4 if Octant 1 in different mod intervals
		sn8 = 0  # This will give Count of rank -4 if Octant 1 in different mod interv
		for i in range(m):
			s = mod*i
			for j in range(mod):
				if (j+s < n-1):
					if (oct[j+s] == 1):
							oct1 = oct1+1  # calculating no. of octant 1 in particular a interval
					elif (oct[j+s] == -1):
							oct2 = oct2+1  # calculating no. of octant -1 in particular a interval
					elif (oct[j+s] == 2):
							oct3 = oct3+1  # calculating no. of octant 2 in particular a interval
					elif (oct[j+s] == -2):
							oct4 = oct4+1  # calculating no. of octant -2 in particular a interval
					elif (oct[j+s] == 3):
							oct5 = oct5+1  # calculating no. of octant 3 in particular a interval
					elif (oct[j+s] == -3):
							oct6 = oct6+1   # calculating no. of octant -3 in particular a interval
					elif (oct[j+s] == 4):
							oct7 = oct7+1  # calculating no. of octant 4 in particular a interval
					elif (oct[j+s] == -4):
							oct8 = oct8+1  # calculating no. of octant -4 in particular a interval
			o1.append(oct1)
			o2.append(oct2)
			o3.append(oct3)
			o4.append(oct4)
			o5.append(oct5)
			o6.append(oct6)
			o7.append(oct7)
			o8.append(oct8)
			req = []  # This list is used for storing the rank in a particular interval
			req.append([oct1, 1])
			req.append([oct2, 2])
			req.append([oct3, 3])
			req.append([oct4, 4])
			req.append([oct5, 5])
			req.append([oct6, 6])
			req.append([oct7, 7])
			req.append([oct8, 8])
			req.sort()  # Sorting req with respect to First Coloumn
			# this will give ranks of different octant in a particular interval
			req_ranks = [0, 0, 0, 0, 0, 0, 0, 0]
			req_ranks[req[0][1]-1] = 8  # Giving rank 8 in sorted list
			req_ranks[req[1][1]-1] = 7  # Giving rank 7 in sorted list
			req_ranks[req[2][1]-1] = 6  # Giving rank 6 in sorted list
			req_ranks[req[3][1]-1] = 5  # Giving rank 5 in sorted list
			req_ranks[req[4][1]-1] = 4  # Giving rank 4 in sorted list
			req_ranks[req[5][1]-1] = 3  # Giving rank 3 in sorted list
			req_ranks[req[6][1]-1] = 2  # Giving rank 2 in sorted list
			req_ranks[req[7][1]-1] = 1  # Giving rank 1 in sorted list
			rank.append(req_ranks)
			rank_id_name = []
			rank_id_name = []
			if req_ranks[0] == 1:  # Givng name and id to Rank 1 Octant
					rank_id_name = [1, "Internal outward Interaction"]
					rank1_id_name.append(rank_id_name)
					sn1 = sn1+1
			elif req_ranks[1] == 1:  # Givng name and id to Rank 1 Octant
					rank_id_name = [-1, "External outward Interaction"]
					rank1_id_name.append(rank_id_name)
					sn2 = sn2+1
			elif req_ranks[2] == 1:  # Givng name and id to Rank 1 Octant
					rank_id_name = [2, "External Ejection"]
					rank1_id_name.append(rank_id_name)
					sn3 = sn3+1
			elif req_ranks[3] == 1:  # Givng name and id to Rank 1 Octant
					rank_id_name = [-2, "Internal Ejection"]
					rank1_id_name.append(rank_id_name)
					sn4 = sn4+1
			elif req_ranks[4] == 1:  # Givng name and id to Rank 1 Octant
						rank_id_name = [3, "External inward Interaction"]
						rank1_id_name.append(rank_id_name)
						sn5 = sn5+1
			elif req_ranks[5] == 1:  # Givng name and id to Rank 1 Octant
						rank_id_name = [-3, "Internal inward Interaction"]
						rank1_id_name.append(rank_id_name)
						sn6 = sn6+1
			elif req_ranks[6] == 1:  # Givng name and id to Rank 1 Octant
						rank_id_name = [4, "Internal Sweep"]
						rank1_id_name.append(rank_id_name)
						sn7 = sn7+1
			elif req_ranks[7] == 1:  # Givng name and id to Rank 1 Octant
						rank_id_name = [-4, "External Sweep"]
						rank1_id_name.append(rank_id_name)
						sn8 = sn8+1
			oct1 = 0
			oct2 = 0
			oct3 = 0
			oct4 = 0
			oct5 = 0
			oct6 = 0
			oct7 = 0
			oct8 = 0


		r=[] # it is giving full data that we need to print
		re=[] # it will give the data 2-d matrix of respective interval
		over=[[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8] # This is matirx of overall transition
		for i in range(0,m) :
			p=i*mod
			a1=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a2=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a3=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a4=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a5=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a6=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a7=[0]*8 # These list will store the value each element in 2-d matrix of different intervals
			a8=[0]*8 # These list will store the value each element in 2-d matrix of different intervals 
			for j in range(0,mod):
				if j+p<=n-3:
					if oct[j+p]==1 and oct[j+p+1]==1 :
						a1[0]=a1[0]+1 # This will keep on increasing the value of each transition in different intervals 
						over[0][0]=over[0][0]+1 # This is for overall transitin matrix
					elif oct[j+p]==1 and oct[j+p+1]==-1:
						a1[1]=a1[1]+1  # This will keep on increasing the value of each transition in different intervals  
						over[0][1]=over[0][1]+1 # This is for overall transitin matrix
					elif oct[j+p]==1 and oct[j+p+1]==2:
						a1[2]=a1[2]+1
						over[0][2]=over[0][2]+1
					elif oct[j+p]==1 and oct[j+p+1]==-2:
						a1[3]=a1[3]+1
						over[0][3]=over[0][3]+1
					elif oct[j+p]==1 and oct[j+1+p]==3:
						a1[4]=a1[4]+1 
						over[0][4]=over[0][4]+1
					elif oct[j+p]==1 and oct[j+p+1]==-3:
						a1[5]=a1[5]+1
						over[0][5]=over[0][5]+1
					elif oct[j+p]==1 and oct[j+1+p]==4:
						a1[6]=a1[6]+1
						over[0][6]=over[0][6]+1
					elif oct[j+p]==1 and oct[j+p+1]==-4:
						a1[7]=a1[7]+1
						over[0][7]=over[0][7]+1
					elif oct[j+p]==-1 and oct[j+p+1]==1 :
						a2[0]=a2[0]+1
						over[1][0]=over[1][0]+1
					elif oct[j+p]==-1 and oct[j+p+1]==-1:
						a2[1]=a2[1]+1    
						over[1][1]=over[1][1]+1
					elif oct[j+p]==-1 and oct[j+p+1]==2:
						a2[2]=a2[2]+1
						over[1][2]=over[1][2]+1
					elif oct[j+p]==-1 and oct[j+p+1]==-2:
						a2[3]=a2[3]+1
						over[1][3]=over[1][3]+1
					elif oct[j+p]==-1 and oct[j+p+1]==3:
						a2[4]=a2[4]+1
						over[1][4]=over[1][4]+1 
					elif oct[j+p]==-1 and oct[j+1+p]==-3:
						a2[5]=a2[5]+1
						over[1][5]=over[1][5]+1
					elif oct[j+p]==-1 and oct[j+p+1]==4:
						a2[6]=a2[6]+1
						over[1][6]=over[1][6]+1
					elif oct[j+p]==-1 and oct[j+1+p]==-4:
						a2[7]=a2[7]+1
						over[1][7]=over[1][7]+1
					elif oct[j+p]==2 and oct[j+p+1]==1 :
						a3[0]=a3[0]+1
						over[2][0]=over[2][0]+1
					elif oct[j+p]==2 and oct[j+p+1]==-1:
						a3[1]=a3[1]+1
						over[2][1]=over[2][1]+1    
					elif oct[j+p]==2 and oct[j+p+1]==2:
						a3[2]=a3[2]+1
						over[2][2]=over[2][2]+1
					elif oct[j+p]==2 and oct[j+p+1]==-2:
						a3[3]=a3[3]+1
						over[2][3]=over[2][3]+1
					elif oct[j+p]==2 and oct[j+p+1]==3:
						a3[4]=a3[4]+1
						over[2][4]=over[2][4]+1 
					elif oct[j+p]==2 and oct[j+1+p]==-3:
						a3[5]=a3[5]+1
						over[2][5]=over[2][5]+1
					elif oct[j+p]==2 and oct[j+p+1]==4:
						a3[6]=a3[6]+1
						over[2][6]=over[2][6]+1
					elif oct[j+p]==2 and oct[j+1+p]==-4:
						a3[7]=a3[7]+1
						over[2][7]=over[2][7]+1
					elif oct[j+p]==-2 and oct[j+p+1]==1 :
						a4[0]=a4[0]+1
						over[3][0]=over[3][0]+1
					elif oct[j+p]==-2 and oct[j+p+1]==-1:
						a4[1]=a4[1]+1
						over[3][1]=over[3][1]+1    
					elif oct[j+p]==-2 and oct[j+p+1]==2:
						a4[2]=a4[2]+1
						over[3][2]=over[3][2]+1
					elif oct[j+p]==-2 and oct[j+p+1]==-2:
						a4[3]=a4[3]+1
						over[3][3]=over[3][3]+1
					elif oct[j+p]==-2 and oct[j+p+1]==3:
						a4[4]=a4[4]+1 
						over[3][4]=over[3][4]+1
					elif oct[j+p]==-2 and oct[j+1+p]==-3:
						a4[5]=a4[5]+1
						over[3][5]=over[3][5]+1
					elif oct[j+p]==-2 and oct[j+p+1]==4:
						a4[6]=a4[6]+1
						over[3][6]=over[3][6]+1
					elif oct[j+p]==-2 and oct[j+1+p]==-4:
						a4[7]=a4[7]+1
						over[3][7]=over[3][7]+1  
					elif oct[j+p]==3 and oct[j+p+1]==1 :
						a5[0]=a5[0]+1
						over[4][0]=over[4][0]+1
					elif oct[j+p]==3 and oct[j+p+1]==-1:
						a5[1]=a5[1]+1
						over[4][1]=over[4][1]+1  
					elif oct[j+p]==3 and oct[j+p+1]==2:
						a5[2]=a5[2]+1
						over[4][2]=over[4][2]+1
					elif oct[j+p]==3 and oct[j+p+1]==-2:
						a5[3]=a5[3]+1
						over[4][3]=over[4][3]+1
					elif oct[j+p]==3 and oct[j+p+1]==3:
						a5[4]=a5[4]+1
						over[4][4]=over[4][4]+1
					elif oct[j+p]==3 and oct[j+1+p]==-3:
						a5[5]=a5[5]+1
						over[4][5]=over[4][5]+1
					elif oct[j+p]==3 and oct[j+p+1]==4:
						a5[6]=a5[6]+1
						over[4][6]=over[4][6]+1
					elif oct[j+p]==3 and oct[j+1+p]==-4:
						a5[7]=a5[7]+1
						over[4][7]=over[4][7]+1 
					elif oct[j+p]==-3 and oct[j+p+1]==1 :
						a6[0]=a6[0]+1
						over[5][0]=over[5][0]+1
					elif oct[j+p]==-3 and oct[j+p+1]==-1:
						a6[1]=a6[1]+1    
						over[5][1]=over[5][1]+1
					elif oct[j+p]==-3 and oct[j+p+1]==2:
						a6[2]=a6[2]+1
						over[5][2]=over[5][2]+1
					elif oct[j+p]==-3 and oct[j+p+1]==-2:
						a6[3]=a6[3]+1
						over[5][3]=over[5][3]+1
					elif oct[j+p]==-3 and oct[j+p+1]==3:
						a6[4]=a6[4]+1
						over[5][4]=over[5][4]+1 
					elif oct[j+p]==-3 and oct[j+1+p]==-3:
						a6[5]=a6[5]+1
						over[5][5]=over[5][5]+1
					elif oct[j+p]==-3 and oct[j+p+1]==4:
						a6[6]=a6[6]+1
						over[5][6]=over[5][6]+1
					elif oct[j+p]==-3 and oct[j+1+p]==-4:
						a6[7]=a6[7]+1
						over[5][7]=over[5][7]+1
					elif oct[j+p]==4 and oct[j+p+1]==1 :
						a7[0]=a7[0]+1
						over[6][0]=over[6][0]+1
					elif oct[j+p]==4 and oct[j+p+1]==-1:
						a7[1]=a7[1]+1
						over[6][1]=over[6][1]+1
					elif oct[j+p]==4 and oct[j+p+1]==2:
						a7[2]=a7[2]+1
						over[6][2]=over[6][2]+1
					elif oct[j+p]==4 and oct[j+p+1]==-2:
						a7[3]=a7[3]+1
						over[6][3]=over[6][3]+1
					elif oct[j+p]==4 and oct[j+p+1]==3:
						a7[4]=a7[4]+1
						over[6][4]=over[6][4]+1
					elif oct[j+p]==4 and oct[j+1+p]==-3:
						a7[5]=a7[5]+1
						over[6][5]=over[6][5]+1
					elif oct[j+p]==4 and oct[j+p+1]==4:
						a7[6]=a7[6]+1
						over[6][6]=over[6][6]+1
					elif oct[j+p]==4 and oct[j+1+p]==-4:
						a7[7]=a7[7]+1
						over[6][7]=over[6][7]+1
					elif oct[j+p]==-4 and oct[j+p+1]==1 :
						a8[0]=a8[0]+1
						over[7][0]=over[7][0]+1
					elif oct[j+p]==-4 and oct[j+p+1]==-1:
						a8[1]=a8[1]+1 
						over[7][1]=over[7][1]+1
					elif oct[j+p]==-4 and oct[j+p+1]==2:
						a8[2]=a8[2]+1
						over[7][2]=over[7][2]+1
					elif oct[j+p]==-4 and oct[j+p+1]==-2:
						a8[3]=a8[3]+1
						over[7][3]=over[7][3]+1
					elif oct[j+p]==-4 and oct[j+p+1]==3:
						a8[4]=a8[4]+1
						over[7][4]=over[7][4]+1
					elif oct[j+p]==-4 and oct[j+1+p]==-3:
						a8[5]=a8[5]+1
						over[7][5]=over[7][5]+1
					elif oct[j+p]==-4 and oct[j+p+1]==4:
						a8[6]=a8[6]+1
						over[7][6]=over[7][6]+1
					elif oct[j+p]==-4 and oct[j+1+p]==-4:
						a8[7]=a8[7]+1
						over[7][7]=over[7][7]+1
			if i==0 :
				r.append(over)
			re.append(a1)  # Making  a 2-d Matrix req by appending eight 1-d matrix
			re.append(a2) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a3) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a4) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a5) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a6) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a7) # Making a 2-d Matrix req by appending eight 1-d matrix
			re.append(a8) # Making a 2-d Matrix req by appending eight 1-d matrix
			c=re.copy() 
			r.append(c) # appending the req matrix in r
			re.clear() # clearing req matrix so that new data may get store into it
		h = ["+1","-1","+2","-2","+3","-3","+4","-4"]
		row=[]
		j=0
		a=0
		b=0
		for x in range(n-2):
			if x==0:
				continue
			if(x==1):
				s="mod "+str(mod)		
			elif(x>=2 and x<2+m):
				if(x==1+m):# it will work only if x=1+m
					z=j*mod
					y=(j+1)*mod
					s=str(z)+"-"+str(n-2)
					j=j+1
				else:
					z=j*mod
					y=(j+1)*mod-1
					s=str(z)+"-"+str(y)
					j=j+1
			elif (x-(2+m))%9==0 and x<9*(m+1)+2+m:
				if x ==2+m:
					row.append(["","Overall transition Count","","","","","","","",""])
					row.append(["","To","","","","","","","",""])
					row.append(["","Octant#","+1","-1","+2","-2","+3","-3","+4","-4"])
				else:
					row.append(["","","","","","","","","",""])
					row.append(["","","","","","","","","",""])
					row.append(["","Mod transition Count","","","","","","","",""])
					row.append(["",str(((x-(2+m))//9-1)*mod)+"-"+str(np.minimum(((x-(2+m))//9)*mod-1,n-2)),"To","","","","","","",""])
					row.append(["","Octant#","+1","-1","+2","-2","+3","-3","+4","-4"])
			elif x<9*(m+1)+(2+m):
				if (x-(2+m))%9==1:
					row.append(["From",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])	
				else:
					row.append(["",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])
				a=a+1
				if a==8:
					a=0
					b=b+1
			else:
				row.append(["","","","","","","","","",""])
		z=len(row)
			# writing of data in sheet1
		os.chdir(curr)
		os.chdir("output")
		from openpyxl import Workbook
		book = Workbook()
		sheet = book.active
		rows = []
		rows.append(["", '', '', '', '', '', '', "", "", "", "", "", "", "Overall Octant count", "", "", "", "", "", "", "","", "", "", "", "", "", "", "", "", "","", '', '', row[0][0],row[0][1],row[0][2],row[0][3],row[0][4],row[0][5],row[0][6],row[0][7],row[0][8],row[0][9], "Longest Subsquence Length", "", "", "", "Longest Subsquence Length with Range", "", ""])
        
		rows.append(["T", 'U', 'V', 'W', 'Uavg', 'Vavg', 'Wavg', "U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg", "Octant", "", "", "", '', '', '', "", "", "", "", "", "", '', '', '', "", "", "", "", "", "","", row[1][0],row[1][1],row[1][2],row[1][3],row[1][4],row[1][5],row[1][6],row[1][7],row[1][8],row[1][9]])  # Printing of heading row
		j = 0
		a = 0
		b = 0
		for x in range(n-2):
			if x>1000:
				rows.append([time1[x], u1[x], v1[x], w1[x],"","","", u2[x], v2[x], w2[x], oct[x]])
			elif x == 0:
				rows.append([time1[x], u1[x], v1[x], w1[x], u_mean, v_mean, w_mean, u2[x], v2[x], w2[x], oct[x], "", "", "Octant ID", "1", "-1", "2", "-2", "3", "-3","4","-4","Rank Octant 1","Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant NAME","",row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
							# c8, rank[x][0], rank[x][1], rank[x][2], rank[x][3], rank[x][4], rank[x][5], rank[x][6], rank[x][7], rank1_id_name[x][0], rank1_id_name[x][1]])
			elif (x == 1):
				s = "mod "+str(mod)
				rows.append([time1[x], u1[x], v1[x], w1[x], "", '', '', u2[x], v2[x], w2[x], oct[x], "", s, "Overall count", c1, c2, c3, c4, c5, c6,c7,c8,rank[x][0], rank[x][1], rank[x][2], rank[x][3], rank[x][4], rank[x][5], rank[x][6], rank[x][7], rank1_id_name[x][0], rank1_id_name[x][1],"",row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif (x >= 2 and x < 2+m):
				if (x == 1+m):  # it will work only if x=1+m
					z = j*mod
					y = (j+1)*mod
					s = str(z)+"-"+str(n-2)
					rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x], oct[x], "","", s, o1[x-2], o2[x-2], o3[x-2], o4[x-2], o5[x-2], o6[x-2], o7[x-2], o8[x-2],
								rank[x-1][0], rank[x-1][1], rank[x-1][2], rank[x-1][3], rank[x-1][4], rank[x-1][5], rank[x-1][6], rank[x-1][7], rank1_id_name[x-1][0], rank1_id_name[x-1][1],"",row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
					j = j+1
				else:
					z = j*mod
					y = (j+1)*mod-1
					s = str(z)+"-"+str(y)
					rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x], oct[x], "","", s, o1[x-2], o2[x-2], o3[x-2], o4[x-2], o5[x-2], o6[x-2], o7[x-2], o8[x-2],
								rank[x-1][0], rank[x-1][1], rank[x-1][2], rank[x-1][3], rank[x-1][4], rank[x-1][5], rank[x-1][6], rank[x-1][7], rank1_id_name[x-1][0], rank1_id_name[x-1][1],"",row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
					j = j+1
			elif x == m+3:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x], oct[x],
							"", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "", "", "Octant ID", "Octant Name", "Count of Rank 1 Mod values", "", "",row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+4:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x],
							oct[x],"", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","", "1", "Internal Outward Interaction", sn1, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+5:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x],
							oct[x],"", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","","-1", "External Outward Interaction", sn2, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+6:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x],
							w2[x], oct[x], "", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "", "", "2", "External Ejection", sn3, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+7:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x],
							w2[x], oct[x],"", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","", "-2", "Internal Ejection", sn4, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+8:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x],
							oct[x],"", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","" ,"3", "External inward Interaction", sn5, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+9:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x], w2[x],
							oct[x], "", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "", "","-3", "Internal inward Interaction", sn6, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+10:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x],
							w2[x], oct[x], "", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","", "4", "Internal Sweep", sn7, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			elif x == m+11:  # Printing the table of count of diffent rank 1 octants
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x], v2[x],
							w2[x], oct[x], "", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","", "-4", "External Sweep", sn8, "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			else:  # Printing  the else data
				rows.append([time1[x], u1[x], v1[x], w1[x], "", "", "", u2[x],
							v2[x], w2[x], oct[x],  "", "", "", "", "", "" ,"", "", "", "" ,"", "", "", "", "", "","", "", "", "", "", "", row[x+2][0],row[x+2][1],row[x+2][2],row[x+2][3],row[x+2][4],row[x+2][5],row[x+2][6],row[x+2][7],row[x+2][8],row[x+2][9],"",tu3[x][0],tu3[x][1],tu3[x][2],"",e1[x],b1[x],d1[x]])
			
		for i in rows:
			sheet.append(i)

		maan = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
		sheet.conditional_formatting.add(f"W4:AE{2+(n-2)//mod+4}", CellIsRule(operator='equal', formula=[1], fill=maan))
		thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
		p=4
		f=len(e1)-1001
		for x in range(2+(n-2)//mod):
			for y in range(35,44):
				sheet.cell(row=p-1, column=y).border = thin_border
			for z in range(8):
				for y in range(35,44):
					sheet.cell(row=p, column=y).border = thin_border
				p=p+1
			p=p+5
		for x in range(3,6+(n-2)//mod):
			for z in range(14,33):
				sheet.cell(row=x, column=z).border = thin_border
		for x in range(7+(n-2)//mod,16+(n-2)//mod):
			for z in range(29,32):
				sheet.cell(row=x, column=z).border = thin_border
		for x in range(3,12):
			for z in range(45,48):
				sheet.cell(row=x, column=z).border = thin_border	
		for x in range(3,f+4):
			for z in range(49,52):
				sheet.cell(row=x, column=z).border = thin_border
		book.save(list[:-5]+" cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
		os.chdir(curr)
		   

mod = 5000
octant_analysis(mod)


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
			
